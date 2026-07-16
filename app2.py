from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
import sqlite3
from datetime import date
import pandas as pd
from io import BytesIO
import io
from flask import send_file
from PyPDF2 import PdfReader
import re
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

app = Flask(__name__)
app.secret_key = "secret-nota"

DB_PATH = "nota_belanja.db"

# === DB SETUP ===
def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS NotaBelanja (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tanggal TEXT NOT NULL,
                supplier TEXT NOT NULL,
                nomor_faktur TEXT UNIQUE NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS ItemBelanja (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nota_id INTEGER,
                nama_barang TEXT NOT NULL,
                jumlah REAL,
                satuan TEXT,
                harga_satuan REAL,
                total_harga REAL,
                FOREIGN KEY(nota_id) REFERENCES NotaBelanja(id)
            )
        """)

                # === Tambahan untuk Penjualan ===
        cur.execute("""
            CREATE TABLE IF NOT EXISTS Penjualan (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tanggal_jual TEXT NOT NULL,
                konsumen TEXT NOT NULL,
                nomor_faktur_beli TEXT NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS ItemPenjualan (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                penjualan_id INTEGER,
                nama_barang TEXT,
                jumlah REAL,
                satuan TEXT,
                harga_beli REAL,
                harga_jual REAL,
                total_jual REAL,
                FOREIGN KEY(penjualan_id) REFERENCES Penjualan(id)
            )
        """)

                # === Tabel pengaturan layout faktur ===
        cur.execute("""
            CREATE TABLE IF NOT EXISTS LayoutSetting (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nama_koperasi TEXT,
                alamat TEXT,
                kontak TEXT,
                posisi_header TEXT,
                jarak_header REAL DEFAULT 12,
                font_size_header INTEGER DEFAULT 11
            )
        """)

        conn.commit()
                # Pastikan kolom tambahan untuk layout tahap 2 ada
        try:
            cur.execute("ALTER TABLE LayoutSetting ADD COLUMN table_header_color TEXT DEFAULT 'lightgrey'")
        except:
            pass
        try:
            cur.execute("ALTER TABLE LayoutSetting ADD COLUMN font_size_table INTEGER DEFAULT 9")
        except:
            pass
        try:
            cur.execute("ALTER TABLE LayoutSetting ADD COLUMN font_size_total INTEGER DEFAULT 10")
        except:
            pass
        try:
            cur.execute("ALTER TABLE LayoutSetting ADD COLUMN footer_align TEXT DEFAULT 'CENTER'")
        except:
            pass
        try:
            cur.execute("ALTER TABLE LayoutSetting ADD COLUMN spacing_after_table REAL DEFAULT 10")
        except:
            pass


       

# === HOME FORM ===
@app.route("/", methods=["GET"])
def index():
    return render_template("index.html", today=date.today().strftime("%Y-%m-%d"))

# === HANDLE SUBMIT ===
@app.route("/prepare", methods=["POST"])
def prepare():
    tanggal = request.form.get("tanggal")
    supplier = request.form.get("supplier")
    nomor_faktur = request.form.get("nomor_faktur")
    count = int(request.form.get("count", 0))

    if not tanggal or not supplier or not nomor_faktur:
        flash("⚠️ Mohon lengkapi Tanggal, Supplier, dan Nomor Faktur sebelum menyimpan!", "warning")
        return redirect(url_for("index"))

    items = []
    for i in range(count):
        nama = request.form.get(f"nama_{i}", "").strip()
        jumlah = request.form.get(f"jumlah_{i}", "").strip()
        satuan = request.form.get(f"satuan_{i}", "").strip()
        harga = request.form.get(f"harga_{i}", "").strip()

        if not nama:
            continue

        if not satuan:
            flash(f"⚠️ Baris ke-{i+1}: satuan belum diisi!", "danger")
            return redirect(url_for("index"))

        try:
            jumlah = float(jumlah or 0)
            harga = float(harga or 0)
        except ValueError:
            flash(f"⚠️ Baris ke-{i+1}: jumlah atau harga tidak valid!", "danger")
            return redirect(url_for("index"))

        if harga <= 0:
            flash(f"⚠️ Baris ke-{i+1}: harga harus lebih dari 0!", "danger")
            return redirect(url_for("index"))

        total = jumlah * harga
        items.append((nama, jumlah, satuan, harga, total))

    if not items:
        flash("⚠️ Tidak ada barang yang diinputkan!", "warning")
        return redirect(url_for("index"))

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    try:
        # 🧹 Hapus nota lama dengan nomor faktur yang sama (jika ada)
        cur.execute("SELECT id FROM NotaBelanja WHERE nomor_faktur = ?", (nomor_faktur,))
        nota = cur.fetchone()
        if nota:
            nota_id = nota[0]
            cur.execute("DELETE FROM ItemBelanja WHERE nota_id = ?", (nota_id,))
            cur.execute("DELETE FROM NotaBelanja WHERE id = ?", (nota_id,))

        # 🆕 Buat nota baru
        cur.execute("""
            INSERT INTO NotaBelanja (tanggal, supplier, nomor_faktur)
            VALUES (?, ?, ?)
        """, (tanggal, supplier, nomor_faktur))
        nota_id = cur.lastrowid

        # 💾 Simpan item baru
        for nama, jumlah, satuan, harga, total in items:
            cur.execute("""
                INSERT INTO ItemBelanja (nota_id, nama_barang, jumlah, satuan, harga_satuan, total_harga)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (nota_id, nama, jumlah, satuan, harga, total))

        conn.commit()
        flash(f"✅ Nota {nomor_faktur} berhasil disimpan ulang ({len(items)} item).", "success")

    except Exception as e:
        conn.rollback()
        flash(f"❌ Terjadi kesalahan saat menyimpan: {e}", "danger")

    finally:
        conn.close()

    return redirect(url_for("index"))



# === AUTOCOMPLETE BARANG ===
@app.route("/autocomplete", methods=["GET"])
def autocomplete():
    keyword = request.args.get("term", "").strip().lower()
    if not keyword:
        return jsonify(results=[])

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT nama_barang 
        FROM ItemBelanja 
        WHERE LOWER(nama_barang) LIKE ?
        ORDER BY nama_barang LIMIT 10
    """, (f"%{keyword}%",))
    results = [row["nama_barang"] for row in cur.fetchall()]
    conn.close()
    return jsonify(results=results)

@app.route("/layout_wizard", methods=["GET", "POST"])
def layout_wizard():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if request.method == "POST":
        nama = request.form["nama_koperasi"]
        alamat = request.form["alamat"]
        kontak = request.form["kontak"]
        posisi = request.form["posisi_header"]
        jarak = request.form["jarak_header"]
        font_size = request.form["font_size_header"]
        table_color = request.form.get("table_header_color", "lightgrey")
        font_table = request.form.get("font_size_table", 9)
        font_total = request.form.get("font_size_total", 10)
        footer_align = request.form.get("footer_align", "CENTER")
        spacing_after = request.form.get("spacing_after_table", 10)

        cur.execute("DELETE FROM LayoutSetting")
        cur.execute("""
            INSERT INTO LayoutSetting (
                nama_koperasi, alamat, kontak, posisi_header, jarak_header, font_size_header,
                table_header_color, font_size_table, font_size_total, footer_align, spacing_after_table
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (nama, alamat, kontak, posisi, jarak, font_size, table_color, font_table, font_total, footer_align, spacing_after))
        conn.commit()
        conn.close()


        flash("✅ Layout faktur berhasil disimpan!", "success")
        return redirect("/layout_wizard")

    # Ambil data setting terakhir
    cur.execute("SELECT * FROM LayoutSetting LIMIT 1")
    setting = cur.fetchone()
    conn.close()
    return render_template("layout_wizard.html", setting=setting)

# === REPORT ===
@app.route("/report")
def report():
    tanggal = request.args.get("tanggal", date.today().strftime("%Y-%m-%d"))
    nomor_faktur = request.args.get("nomor_faktur", "").strip()

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    query = """
        SELECT n.nomor_faktur, n.tanggal, i.nama_barang, i.jumlah, i.satuan, 
               i.harga_satuan, i.total_harga
        FROM NotaBelanja n
        JOIN ItemBelanja i ON n.id = i.nota_id
        WHERE 1=1
    """
    params = []

    if nomor_faktur:
        query += " AND n.nomor_faktur LIKE ?"
        params.append(f"%{nomor_faktur}%")
    else:
        query += " AND n.tanggal = ?"
        params.append(tanggal)

    query += " ORDER BY n.nomor_faktur ASC"

    cur.execute(query, params)
    rows = cur.fetchall()
    total = sum(r["total_harga"] for r in rows) if rows else 0
    conn.close()

    return render_template("report.html",
                           tanggal=tanggal,
                           rows=rows,
                           total=total,
                           nomor_faktur=nomor_faktur)

# === FORM PENJUALAN ===
@app.route("/penjualan", methods=["GET", "POST"])
def penjualan():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT nomor_faktur FROM NotaBelanja ORDER BY nomor_faktur DESC")
    nota_list = [r["nomor_faktur"] for r in cur.fetchall()]
    conn.close()

    if request.method == "POST":
        tanggal_jual = request.form.get("tanggal_jual")
        konsumen = request.form.get("konsumen")
        nomor_faktur_beli = request.form.get("nomor_faktur_beli")
        count = int(request.form.get("count", 0))

        if not (tanggal_jual and konsumen and nomor_faktur_beli):
            flash("⚠️ Lengkapi semua kolom terlebih dahulu!", "warning")
            return redirect(url_for("penjualan"))

        items = []
        for i in range(count):
            nama = request.form.get(f"nama_{i}")
            jumlah = request.form.get(f"jumlah_{i}")
            satuan = request.form.get(f"satuan_{i}")
            harga_beli = request.form.get(f"harga_beli_{i}")
            harga_jual = request.form.get(f"harga_jual_{i}")

            if not nama or not harga_jual:
                continue

            try:
                jumlah = float(jumlah)
                harga_beli = float(harga_beli)
                harga_jual = float(harga_jual)
            except ValueError:
                continue

            total_jual = jumlah * harga_jual
            items.append((nama, jumlah, satuan, harga_beli, harga_jual, total_jual))

        if not items:
            flash("⚠️ Tidak ada barang yang diinputkan!", "warning")
            return redirect(url_for("penjualan"))

        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO Penjualan (tanggal_jual, konsumen, nomor_faktur_beli)
            VALUES (?, ?, ?)
        """, (tanggal_jual, konsumen, nomor_faktur_beli))
        penjualan_id = cur.lastrowid

        for nama, jumlah, satuan, harga_beli, harga_jual, total_jual in items:
            cur.execute("""
                INSERT INTO ItemPenjualan (penjualan_id, nama_barang, jumlah, satuan, harga_beli, harga_jual, total_jual)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (penjualan_id, nama, jumlah, satuan, harga_beli, harga_jual, total_jual))

        conn.commit()
        conn.close()

        flash(f"✅ Penjualan ke {konsumen} berhasil disimpan.", "success")
        return redirect(url_for("penjualan"))

    return render_template("penjualan.html", nota_list=nota_list, today=date.today().strftime("%Y-%m-%d"))


# === API: Ambil Detail Nota Pembelian ===
@app.route("/get_items_by_nota")
def get_items_by_nota():
    nota = request.args.get("nota", "").strip()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT i.nama_barang, i.jumlah, i.satuan, i.harga_satuan
        FROM NotaBelanja n
        JOIN ItemBelanja i ON n.id = i.nota_id
        WHERE n.nomor_faktur = ?
    """, (nota,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify(rows)


# === SIMPAN PENJUALAN ===
@app.route("/simpan_penjualan", methods=["POST"])
def simpan_penjualan():
    tanggal = request.form.get("tanggal")
    konsumen = request.form.get("konsumen")
    nota_beli = request.form.get("nota_beli")

    if not tanggal or not konsumen or not nota_beli:
        flash("⚠️ Lengkapi semua kolom terlebih dahulu!", "warning")
        return redirect("/penjualan")

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # Pastikan tabel ada
    cur.execute("""
        CREATE TABLE IF NOT EXISTS Penjualan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tanggal_jual TEXT NOT NULL,
            konsumen TEXT NOT NULL,
            nomor_faktur_beli TEXT UNIQUE NOT NULL
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS ItemPenjualan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            penjualan_id INTEGER,
            nama_barang TEXT,
            jumlah REAL,
            satuan TEXT,
            harga_beli REAL,
            harga_jual REAL,
            total_jual REAL,
            FOREIGN KEY(penjualan_id) REFERENCES Penjualan(id)
        )
    """)
    conn.commit()

    # Cek apakah faktur sudah ada
    cur.execute("SELECT id FROM Penjualan WHERE nomor_faktur_beli = ?", (nota_beli,))
    row = cur.fetchone()

    if row:
        penjualan_id = row[0]
        # Hapus semua item lama untuk faktur ini
        cur.execute("DELETE FROM ItemPenjualan WHERE penjualan_id = ?", (penjualan_id,))
        mode = "update"
    else:
        # Faktur baru
        cur.execute("""
            INSERT INTO Penjualan (tanggal_jual, konsumen, nomor_faktur_beli)
            VALUES (?, ?, ?)
        """, (tanggal, konsumen, nota_beli))
        penjualan_id = cur.lastrowid
        mode = "new"

    # Simpan item baru
    count = 0
    for key in request.form:
        if key.startswith("nama_"):
            index = key.split("_")[1]
            nama = request.form.get(f"nama_{index}", "").strip()
            jumlah = float(request.form.get(f"jumlah_{index}", 0) or 0)
            satuan = request.form.get(f"satuan_{index}", "").strip()
            harga_beli = float(request.form.get(f"harga_beli_{index}", 0) or 0)
            harga_jual = float(request.form.get(f"harga_jual_{index}", 0) or 0)
            total = jumlah * harga_jual

            if not nama or harga_jual <= 0:
                continue

            cur.execute("""
                INSERT INTO ItemPenjualan
                (penjualan_id, nama_barang, jumlah, satuan, harga_beli, harga_jual, total_jual)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (penjualan_id, nama, jumlah, satuan, harga_beli, harga_jual, total))
            count += 1

    conn.commit()
    conn.close()

    # Pesan flash
    if mode == "update":
        flash(f"✅ Faktur {nota_beli} diperbarui ({count} item disimpan, data lama dihapus).", "success")
    else:
        flash(f"✅ Penjualan baru disimpan ({count} item) untuk faktur {nota_beli}.", "success")

    return redirect("/penjualan")


@app.route("/cek_faktur_penjualan")
def cek_faktur_penjualan():
    nomor_faktur = request.args.get("nomor_faktur")
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT id FROM Penjualan WHERE nomor_faktur_beli = ?", (nomor_faktur,))
    ada = cur.fetchone()
    conn.close()
    return jsonify({"exists": bool(ada)})

# === API: Gabungkan ItemBelanja + ItemPenjualan berdasarkan Nomor Faktur ===
@app.route('/cari_faktur_penjualan')
def cari_faktur_penjualan():
    nomor_faktur = request.args.get('nomor_faktur')
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Ambil data nota pembelian
    cur.execute("""
        SELECT tanggal, supplier AS konsumen
        FROM NotaBelanja
        WHERE nomor_faktur = ?
    """, (nomor_faktur,))
    nota = cur.fetchone()

    if not nota:
        return jsonify({"status": "not_found"})

    # Ambil data item pembelian
    cur.execute("""
        SELECT nama_barang, jumlah, satuan, harga_satuan AS harga_beli
        FROM ItemBelanja
        WHERE nota_id = (SELECT id FROM NotaBelanja WHERE nomor_faktur = ?)
    """, (nomor_faktur,))
    items = [dict(row) for row in cur.fetchall()]

    # Cek apakah sudah pernah dijual
    for item in items:
        cur.execute("""
            SELECT harga_jual
            FROM ItemPenjualan
            WHERE nama_barang = ?
            ORDER BY id DESC LIMIT 1
        """, (item['nama_barang'],))
        jual = cur.fetchone()
        item['harga_jual'] = jual['harga_jual'] if jual else None

    conn.close()
    return jsonify({
        "status": "ok",
        "tanggal": nota["tanggal"],
        "konsumen": nota["konsumen"],
        "items": items
    })

@app.route("/suggest_harga", methods=["POST"])
def suggest_harga():
    data = request.get_json()
    items = data.get("items", [])

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    hasil = []
    for item in items:
        nama = item.get("nama")
        satuan = item.get("satuan")

        # Ambil harga jual terakhir
        cur.execute("""
            SELECT harga_jual FROM ItemPenjualan
            WHERE nama_barang = ? COLLATE NOCASE AND satuan = ? COLLATE NOCASE
            ORDER BY id DESC LIMIT 1
        """, (nama, satuan))

        row = cur.fetchone()

        if row and row[0]:
            hasil.append(row[0])
        else:
            # fallback: ambil harga beli + 20%
            cur.execute("""
                SELECT harga_beli FROM ItemPenjualan
                WHERE nama_barang = ? COLLATE NOCASE AND satuan = ? COLLATE NOCASE
                ORDER BY id DESC LIMIT 1
            """, (nama, satuan))

            hb = cur.fetchone()
            hasil.append(round(hb[0] * 1.2) if hb else None)

    conn.close()
    return jsonify({"status": "ok", "saran": hasil})


@app.route("/export_penjualan_excel")
def export_penjualan_excel():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Ambil data gabungan Penjualan + ItemPenjualan
    query = """
        SELECT 
            p.tanggal_jual AS Tanggal,
            p.nomor_faktur_beli AS Nomor_Faktur,
            p.konsumen AS Konsumen,
            i.nama_barang AS Nama_Barang,
            i.jumlah AS Jumlah,
            i.satuan AS Satuan,
            i.harga_beli AS Harga_Beli,
            i.harga_jual AS Harga_Jual,
            i.total_jual AS Total_Jual
        FROM Penjualan p
        JOIN ItemPenjualan i ON p.id = i.penjualan_id
        ORDER BY p.tanggal_jual DESC, p.nomor_faktur_beli ASC
    """
    cur.execute(query)
    rows = cur.fetchall()
    conn.close()

    if not rows:
        flash("⚠️ Tidak ada data penjualan untuk diekspor.", "warning")
        return redirect("/penjualan")

    # Buat dataframe dengan urutan kolom yang rapi
    df = pd.DataFrame(rows, columns=[
        "Tanggal", "Nomor_Faktur", "Konsumen",
        "Nama_Barang", "Jumlah", "Satuan",
        "Harga_Beli", "Harga_Jual", "Total_Jual"
    ])

    # Format angka
    df["Jumlah"] = df["Jumlah"].astype(float)
    df["Harga_Beli"] = df["Harga_Beli"].astype(float)
    df["Harga_Jual"] = df["Harga_Jual"].astype(float)
    df["Total_Jual"] = df["Total_Jual"].astype(float)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="Data Penjualan", index=False)
        workbook = writer.book
        worksheet = writer.sheets["Data Penjualan"]

        # Format Header
        header_fmt = workbook.add_format({
            "bold": True,
            "align": "center",
            "valign": "vcenter",
            "fg_color": "#B7DEE8",
            "border": 1
        })

        # Format isi angka
        money_fmt = workbook.add_format({
            "num_format": '#,##0',
            "align": "right",
            "border": 1
        })

        # Format isi teks
        text_fmt = workbook.add_format({
            "align": "left",
            "border": 1
        })

        # Terapkan format header dan lebar kolom
        for col_num, col_name in enumerate(df.columns):
            worksheet.write(0, col_num, col_name, header_fmt)

            if col_name in ["Jumlah", "Harga_Beli", "Harga_Jual", "Total_Jual"]:
                worksheet.set_column(col_num, col_num, 15, money_fmt)
            elif col_name in ["Tanggal", "Nomor_Faktur"]:
                worksheet.set_column(col_num, col_num, 18, text_fmt)
            else:
                worksheet.set_column(col_num, col_num, 20, text_fmt)

        # Tambahkan autofilter di header
        worksheet.autofilter(0, 0, len(df), len(df.columns) - 1)

        # Tambahkan total di bawah tabel
        total_row = len(df) + 1
        worksheet.write(total_row, 7, "TOTAL :", header_fmt)
        worksheet.write_formula(total_row, 8, f"=SUM(I2:I{total_row})", money_fmt)

    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="Data_Penjualan_Rapi.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/export_excel")
def export_excel():
    tanggal = request.args.get("tanggal", date.today().strftime("%Y-%m-%d"))
    nomor_faktur = request.args.get("nomor_faktur", "").strip()

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    query = """
        SELECT n.nomor_faktur, n.tanggal, i.nama_barang, i.jumlah, i.satuan,
               i.harga_satuan, i.total_harga
        FROM NotaBelanja n
        JOIN ItemBelanja i ON n.id = i.nota_id
        WHERE 1=1
    """
    params = []

    if nomor_faktur:
        query += " AND n.nomor_faktur LIKE ?"
        params.append(f"%{nomor_faktur}%")
    else:
        query += " AND n.tanggal = ?"
        params.append(tanggal)

    query += " ORDER BY n.nomor_faktur ASC"

    cur.execute(query, params)
    rows = cur.fetchall()
    conn.close()

    if not rows:
        flash("Tidak ada data untuk diexport", "warning")
        return redirect("/report")

    df = pd.DataFrame(rows)
    path = "report_harian.xlsx"
    df.to_excel(path, index=False)

    return send_file(path, as_attachment=True)

def get_most_frequent_price(nama_barang):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        SELECT harga_satuan, COUNT(*) AS freq
        FROM ItemBelanja
        WHERE LOWER(nama_barang) = LOWER(?)
        GROUP BY harga_satuan
        ORDER BY freq DESC
        LIMIT 1
    """, (nama_barang,))
    result = cur.fetchone()
    conn.close()
    return result[0] if result else None

@app.route('/cari_barang')
def cari_barang():
    from fuzzywuzzy import process, fuzz  # atau rapidfuzz fallback
    q = request.args.get('q', '').strip().lower()

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Ambil seluruh data (tanpa DISTINCT)
    cur.execute("""
        SELECT id, nama_barang, satuan, harga_satuan
        FROM ItemBelanja
        WHERE nama_barang IS NOT NULL AND nama_barang <> ''
        ORDER BY id DESC
    """)
    rows = cur.fetchall()
    conn.close()

    if not q:
        return jsonify(results=[])

    # Semua nama barang
    all_names = [d['nama_barang'] for d in rows]

    # Fuzzy matching
    hasil = process.extract(q, all_names, limit=30, scorer=fuzz.token_set_ratio)

    # Filter fuzzy + wajib mengandung kata pencarian (di awal, tengah, atau akhir)
    hasil_filtered = [
        nama for nama, skor in hasil 
        if skor >= 60 and q in nama.lower()
    ]

    # === UNIQUE BERDASARKAN RECORD TERBARU ===
    unique_latest = {}
    for r in rows:
        nama = r["nama_barang"]
        if nama in hasil_filtered and nama not in unique_latest:
            unique_latest[nama] = r  # record terbaru

    # Format output JSON
    results = []
    for r in unique_latest.values():
        results.append({
            "nama": r["nama_barang"],
            "satuan": r["satuan"],
            "harga": r["harga_satuan"]
        })

    return jsonify(results=results)

@app.route("/get_harga_terbanyak")
def get_harga_terbanyak():
    nama = request.args.get("nama", "").strip()
    if not nama:
        return jsonify({"harga": None})
    harga = get_most_frequent_price(nama)
    return jsonify({"harga": harga})

@app.route("/get_harga_terbaru")
def get_harga_terbaru():
    nama = request.args.get("nama", "").strip()
    if not nama:
        return jsonify({"harga": None})

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.execute("""
        SELECT harga_satuan
        FROM ItemBelanja
        WHERE nama_barang = ?
        ORDER BY tanggal DESC, id DESC
        LIMIT 1
    """, (nama,))

    row = cur.fetchone()
    conn.close()

    return jsonify({
        "harga": row[0] if row else None
    })

@app.route('/cek_fuzzy_barang')
def cek_fuzzy_barang():
    try:
        from fuzzywuzzy import fuzz, process
    except ImportError:
        from rapidfuzz import fuzz, process  # fallback jika fuzzywuzzy tidak ada

    q = request.args.get('q', '').strip()
    if not q:
        return jsonify(match=None, score=0)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT nama_barang FROM ItemBelanja WHERE nama_barang IS NOT NULL AND nama_barang <> ''")
    all_names = [row['nama_barang'] for row in cur.fetchall()]
    conn.close()

    if not all_names:
        return jsonify(match=None, score=0)

    match, score = process.extractOne(q, all_names, scorer=fuzz.token_set_ratio)
    if score >= 85:
        return jsonify(match=match, score=score)
    else:
        return jsonify(match=None, score=score)

@app.route('/saran_barang')
def saran_barang():
    try:
        from fuzzywuzzy import fuzz, process
    except ImportError:
        from rapidfuzz import fuzz, process

    q = request.args.get('q', '').strip()
    if not q:
        return jsonify(suggestions=[])

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT nama_barang FROM ItemBelanja WHERE nama_barang IS NOT NULL AND nama_barang <> ''")
    all_names = [row['nama_barang'] for row in cur.fetchall()]
    conn.close()

    if not all_names:
        return jsonify(suggestions=[])

    hasil = process.extract(q, all_names, limit=5, scorer=fuzz.token_set_ratio)
    hasil_filtered = [h for h in hasil if h[1] >= 70]

    suggestions = [{"nama": h[0], "score": h[1]} for h in hasil_filtered]
    return jsonify(suggestions=suggestions)

@app.route("/export_penjualan_excel_by_faktur")
def export_penjualan_excel_by_faktur():
    import io
    import pandas as pd
    from datetime import datetime

    nomor_faktur = request.args.get("nomor_faktur", "").strip()
    if not nomor_faktur:
        flash("⚠️ Nomor faktur wajib diisi!", "warning")
        return redirect("/penjualan")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # === Header Faktur ===
    cur.execute("SELECT id, konsumen, tanggal_jual FROM Penjualan WHERE nomor_faktur_beli = ?", (nomor_faktur,))
    penjualan = cur.fetchone()
    if not penjualan:
        conn.close()
        flash(f"⚠️ Faktur {nomor_faktur} tidak ditemukan.", "warning")
        return redirect("/penjualan")

    # === Data Item ===
    cur.execute("""
        SELECT 
            nama_barang AS 'Nama Barang',
            jumlah AS 'Jumlah',
            satuan AS 'Satuan',
            harga_beli AS 'Harga Beli',
            (jumlah * harga_beli) AS 'Total Beli',
            harga_jual AS 'Harga Jual',
            total_jual AS 'Subtotal'
        FROM ItemPenjualan
        WHERE penjualan_id = ?
    """, (penjualan["id"],))
    rows = cur.fetchall()
    conn.close()

    if not rows:
        flash("⚠️ Tidak ada item untuk faktur ini.", "warning")
        return redirect("/penjualan")

    df = pd.DataFrame(rows)

    # === Buat file Excel ===
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        workbook = writer.book
        worksheet = workbook.add_worksheet("Penjualan")
        writer.sheets["Penjualan"] = worksheet

        # === FORMAT ===
        title_fmt = workbook.add_format({"bold": True, "font_size": 14, "align": "center"})
        header_fmt = workbook.add_format({
            "bold": True, "align": "center", "valign": "vcenter",
            "bg_color": "#cfe2f3", "border": 1
        })
        text_fmt = workbook.add_format({"border": 1})
        money_fmt = workbook.add_format({"num_format": '#,##0', "align": "right", "border": 1})
        info_fmt = workbook.add_format({"bold": True, "align": "left"})
        total_label_fmt = workbook.add_format({
            "bold": True, "align": "right", "border": 1, "bg_color": "#fce5cd"
        })

        # === JUDUL ===
        worksheet.merge_range("A1:G1", "LAPORAN PENJUALAN PER FAKTUR", title_fmt)

        # === INFO FAKTUR ===
        worksheet.write("A3", "Nomor Faktur:", info_fmt)
        worksheet.write("B3", nomor_faktur)
        worksheet.write("A4", "Tanggal Jual:", info_fmt)
        worksheet.write("B4", penjualan["tanggal_jual"])
        worksheet.write("A5", "Konsumen:", info_fmt)
        worksheet.write("B5", penjualan["konsumen"])

        # === HEADER KOLOM ===
        for col_num, col_name in enumerate(df.columns):
            worksheet.write(6, col_num, col_name, header_fmt)

        # === ISI DATA ===
        for row_num, row_data in enumerate(df.values):
            for col_num, cell_value in enumerate(row_data):
                if df.columns[col_num] in ["Harga Beli", "Total Beli", "Harga Jual", "Subtotal"]:
                    worksheet.write(row_num + 7, col_num, cell_value, money_fmt)
                else:
                    worksheet.write(row_num + 7, col_num, cell_value, text_fmt)

        # === FORMAT KOLOM ===
        worksheet.set_column("A:A", 25)
        worksheet.set_column("B:B", 10)
        worksheet.set_column("C:C", 10)
        worksheet.set_column("D:D", 15)
        worksheet.set_column("E:E", 15)
        worksheet.set_column("F:F", 15)
        worksheet.set_column("G:G", 15)

        # === TOTAL DAN LABA ===
        total_row = len(df) + 7
        worksheet.write(total_row, 3, "TOTAL BELI :", total_label_fmt)
        worksheet.write_formula(total_row, 4, f"=SUM(E8:E{total_row})", money_fmt)

        worksheet.write(total_row + 1, 3, "TOTAL PENJUALAN :", total_label_fmt)
        worksheet.write_formula(total_row + 1, 6, f"=SUM(G8:G{total_row})", money_fmt)

        worksheet.write(total_row + 2, 3, "LABA KOTOR :", total_label_fmt)
        worksheet.write_formula(total_row + 2, 6, f"=G{total_row+2}-E{total_row+1}", money_fmt)

    output.seek(0)
    filename = f"Penjualan_{nomor_faktur}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# === API: Ambil Nota Lengkap (header + item) ===
@app.route("/get_nota_detail")
def get_nota_detail():
    nomor_faktur = request.args.get("nomor_faktur", "").strip()
    if not nomor_faktur:
        return jsonify({"status": "error", "message": "Nomor faktur kosong"})

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT id, tanggal, supplier, nomor_faktur
        FROM NotaBelanja
        WHERE nomor_faktur = ?
    """, (nomor_faktur,))
    nota = cur.fetchone()

    if not nota:
        conn.close()
        return jsonify({"status": "not_found", "message": f"Nota {nomor_faktur} tidak ditemukan"})

    cur.execute("""
        SELECT nama_barang, jumlah, satuan, harga_satuan, total_harga
        FROM ItemBelanja
        WHERE nota_id = ?
        ORDER BY id ASC
    """, (nota["id"],))
    items = [dict(r) for r in cur.fetchall()]
    conn.close()

    return jsonify({
        "status": "ok",
        "tanggal": nota["tanggal"],
        "supplier": nota["supplier"],
        "nomor_faktur": nota["nomor_faktur"],
        "items": items
    })

@app.route("/get_all_nota")
def get_all_nota():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT nomor_faktur, supplier
        FROM NotaBelanja
        ORDER BY tanggal DESC, id DESC
        LIMIT 50
    """)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"status": "ok", "notas": rows})

# === API: Riwayat Harga Barang (gabungan pembelian + penjualan) ===
@app.route("/riwayat_harga")
def riwayat_harga():
    nama = request.args.get("nama", "").strip()
    if not nama:
        return jsonify({"status": "error", "message": "Nama barang kosong"})

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # === BELI: 1 baris per tanggal ===
    cur.execute("""
        SELECT n.tanggal AS tanggal, n.supplier AS pihak,
               i.harga_satuan AS harga_beli, NULL AS harga_jual, 'BELI' AS jenis
        FROM ItemBelanja i
        JOIN NotaBelanja n ON i.nota_id = n.id
        WHERE LOWER(i.nama_barang) = LOWER(?)
        GROUP BY n.tanggal
        ORDER BY n.tanggal DESC
    """, (nama,))
    beli = [dict(r) for r in cur.fetchall()]

    # === JUAL: 1 baris per tanggal ===
    cur.execute("""
        SELECT p.tanggal_jual AS tanggal, p.konsumen AS pihak,
               NULL AS harga_beli, i.harga_jual AS harga_jual, 'JUAL' AS jenis
        FROM ItemPenjualan i
        JOIN Penjualan p ON i.penjualan_id = p.id
        WHERE LOWER(i.nama_barang) = LOWER(?)
        GROUP BY p.tanggal_jual
        ORDER BY p.tanggal_jual DESC
    """, (nama,))
    jual = [dict(r) for r in cur.fetchall()]

    data = beli + jual
    data.sort(key=lambda x: x["tanggal"] or "", reverse=True)

    conn.close()
    return jsonify({"status": "ok", "data": data})


@app.route("/cari_item")
def cari_item_page():
    return render_template("cari_item.html")

@app.route("/simpan_arsip_pdf", methods=["POST"])
def simpan_arsip_pdf():
    try:
        file = request.files.get("file")
        nomor_faktur = request.form.get("nomor_faktur", "").strip()

        if not file:
            return jsonify({"status": "error", "message": "File PDF tidak ditemukan di request."})
        if not nomor_faktur:
            return jsonify({"status": "error", "message": "Nomor faktur kosong."})

        # Pastikan folder arsip ada
        folder_arsip = os.path.join(app.static_folder, "arsip_penjualan")
        os.makedirs(folder_arsip, exist_ok=True)

        # Buat nama file aman (ganti / jadi _)
        safe_name = f"{nomor_faktur.replace('/', '_')}.pdf"
        file_path = os.path.join(folder_arsip, safe_name)

        # Simpan file PDF
        file.save(file_path)
        print(f"✅ Arsip tersimpan: {file_path}")

        return jsonify({"status": "ok", "message": "Arsip berhasil disimpan.", "file_url": f"/static/arsip_penjualan/{safe_name}"})
    except Exception as e:
        print("❌ ERROR simpan_arsip_pdf:", str(e))
        return jsonify({"status": "error", "message": str(e)})

    

import os
from flask import send_from_directory

@app.route("/get_daftar_arsip", methods=["GET"])
def get_daftar_arsip():
    folder = os.path.join(app.root_path, "static", "arsip_penjualan")
    if not os.path.exists(folder):
        return jsonify({"status": "ok", "data": []})

    files = []
    for f in sorted(os.listdir(folder)):
        if f.lower().endswith(".pdf"):
            files.append({
                "nama": f,
                "url": url_for('static', filename=f"arsip_penjualan/{f}")
            })

    return jsonify({"status": "ok", "data": files})


@app.route("/cek_arsip_pdf")
def cek_arsip_pdf():
    nomor_faktur = request.args.get("nomor_faktur", "").strip()
    if not nomor_faktur:
        return jsonify({"status": "error", "message": "Nomor faktur kosong"})

    folder_arsip = os.path.join(app.static_folder, "arsip_penjualan")
    safe_name = f"{nomor_faktur.replace('/', '_')}.pdf"
    file_path = os.path.join(folder_arsip, safe_name)

    if os.path.exists(file_path):
        file_url = f"/static/arsip_penjualan/{safe_name}"
        print("✅ File ditemukan:", file_url)  # Tambahkan ini untuk lihat hasil di console Flask
        return jsonify({"status": "ok", "file_url": file_url})
    else:
        print("❌ Tidak ditemukan:", file_path)
        return jsonify({"status": "not_found"})


@app.route("/baca_pdf_arsip")
def baca_pdf_arsip():
    nomor_faktur = request.args.get("nomor_faktur", "").strip()
    if not nomor_faktur:
        return jsonify({"status": "error", "message": "Nomor faktur kosong"})

    folder_arsip = os.path.join("static", "arsip_penjualan")
    safe_name = f"{nomor_faktur.replace('/', '_')}.pdf"
    file_path = os.path.join(folder_arsip, safe_name)

    if os.path.exists(file_path):
        return jsonify({
            "status": "ok",
            "file_url": f"/{file_path.replace(os.sep, '/')}"
        })
    else:
        return jsonify({"status": "not_found"})


@app.route("/get_arsip_pdf")
def get_arsip_pdf():
    nomor_faktur = request.args.get("nomor_faktur", "").strip()
    if not nomor_faktur:
        return jsonify({"status": "error", "message": "Nomor faktur kosong"})

    pdf_path = os.path.join("arsip", f"{nomor_faktur}.pdf")
    if os.path.exists(pdf_path):
        return jsonify({"status": "ok", "path": "/" + pdf_path.replace("\\", "/")})
    else:
        return jsonify({"status": "not_found"})


@app.route("/hapus_arsip_pdf", methods=["POST"])
def hapus_arsip_pdf():
    nama = request.form.get("nama", "")
    if not nama:
        return jsonify({"status": "error", "message": "Nama file tidak diberikan."})

    arsip_folder = os.path.join("static", "arsip_pdf")
    file_path = os.path.join(arsip_folder, nama)

    if not os.path.exists(file_path):
        return jsonify({"status": "error", "message": "File tidak ditemukan."})

    try:
        os.remove(file_path)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


@app.route("/get_pdf_arsip")
def get_pdf_arsip():
    nomor_faktur = request.args.get("nomor_faktur", "").strip()
    if not nomor_faktur:
        return jsonify({"status": "error", "message": "Nomor faktur kosong"})

    pdf_path = f"static/arsip/{nomor_faktur}.pdf"
    if not os.path.exists(pdf_path):
        return jsonify({"status": "not_found", "message": "Arsip PDF tidak ditemukan."})

    # Baca teks dari PDF (optional)
    try:
        reader = PdfReader(pdf_path)
        pdf_text = ""
        for page in reader.pages:
            pdf_text += page.extract_text() + "\n"
    except Exception as e:
        pdf_text = f"Gagal membaca isi PDF: {e}"

    return jsonify({
        "status": "ok",
        "pdf_url": f"/static/arsip/{nomor_faktur}.pdf",
        "pdf_text": pdf_text
    })

@app.route("/get_item_penjualan")
def get_item_penjualan():
    try:
        nomor_faktur = request.args.get("nomor_faktur", "").strip()
        if not nomor_faktur:
            return jsonify({
                "status": "error",
                "message": "Nomor faktur kosong."
            })

        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        cur.execute("""
            SELECT 
                i.nama_barang,
                i.jumlah,
                i.satuan,
                i.harga_jual,
                i.total_jual
            FROM ItemPenjualan i
            JOIN Penjualan p ON i.penjualan_id = p.id
            WHERE p.nomor_faktur_beli = ?
        """, (nomor_faktur,))
        rows = cur.fetchall()
        conn.close()

        if not rows:
            return jsonify({
                "status": "error",
                "message": f"Tidak ditemukan item untuk nota penjualan dengan nomor faktur {nomor_faktur}."
            })

        data = [dict(row) for row in rows]
        return jsonify({"status": "ok", "data": data})

    except sqlite3.OperationalError as e:
        return jsonify({
            "status": "error",
            "message": f"Kesalahan database: {str(e)}"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Terjadi kesalahan: {str(e)}"
        })


@app.route("/upload_pdf", methods=["POST"])
def upload_pdf():
    from PyPDF2 import PdfReader
    import re, os, tempfile, traceback

    print("=== DEBUG upload_pdf ===")

    # --- 1️⃣ Pastikan file ada ---
    if "file" not in request.files:
        print("🚫 Tidak ada file PDF dalam request.")
        return jsonify({"status": "error", "message": "Tidak ada file PDF yang dikirim."})

    file = request.files["file"]
    nomor_faktur = request.form.get("nomor_faktur", "").strip() or "(tidak diketahui)"
    print(f"📦 Menerima file: {file.filename} | tipe: {file.content_type} | faktur: {nomor_faktur}")

    # --- 2️⃣ Simpan dulu ke file sementara agar bisa dibaca PyPDF2 ---
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name

        if os.path.getsize(tmp_path) == 0:
            print("🚫 File PDF kosong!")
            return jsonify({"status": "error", "message": "File PDF kosong atau gagal diambil dari browser."})
    except Exception as e:
        print("❌ Gagal menyimpan file sementara:", e)
        return jsonify({"status": "error", "message": "Gagal menyimpan file PDF sementara."})

    # --- 3️⃣ Baca isi PDF ---
    try:
        reader = PdfReader(tmp_path)
        print(f"📖 Jumlah halaman PDF: {len(reader.pages)}")

        text = ""
        for page in reader.pages:
            t = page.extract_text() or ""
            # Bersihkan format teks PDF agar bisa dikenali pola regex
            t = re.sub(r"(\d)([A-Za-z])", r"\1 \2", t)
            t = re.sub(r"([a-z])([A-Z])", r"\1 \2", t)
            t = re.sub(r"([A-Za-z])(\d)", r"\1 \2", t)
            t = re.sub(r"\s{2,}", " ", t)
            text += t + "\n"

        if not text.strip():
            print("⚠️ Tidak ada teks terbaca dari PDF.")
            return jsonify({"status": "error", "message": "Tidak ada teks yang bisa dibaca dari file PDF ini."})

    except Exception as e:
        print("❌ ERROR saat membaca PDF:")
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"Gagal membaca file PDF: {str(e)}"})

    # --- 4️⃣ Ambil bagian daftar item antara 'Item' dan 'Info Pembayaran' ---
    match = re.search(r"# ?Item(.+?)Info ?Pembayaran", text, re.DOTALL | re.IGNORECASE)
    if not match:
        print("⚠️ Tidak ditemukan bagian daftar item di PDF.")
        return jsonify({"status": "error", "message": "Bagian daftar item tidak ditemukan di PDF."})

    daftar = match.group(1)

    # --- 5️⃣ Regex parsing baris item (dengan satuan opsional) ---
    pattern = re.compile(
        r"^\s*(\d+)\s+([A-Za-z0-9\s]+?)\s+(?:([A-Za-z]+)\s+)?([\d.,]+)\s+Rp\s*([\d.,]+)\s+Rp\s*([\d.,]+)",
        re.IGNORECASE | re.MULTILINE
    )

    def parse_num(val):
        val = val.replace("Rp", "").replace(" ", "").replace(",", "")
        if val.count(".") > 1:
            val = val.replace(".", "", val.count(".") - 1)
        try:
            return float(val)
        except:
            return 0.0

    items = []
    for m in pattern.finditer(daftar):
        satuan = m.group(3).strip() if m.group(3) else "-"
        items.append({
            "no": int(m.group(1)),
            "nama_barang": m.group(2).strip(),
            "satuan": satuan,
            "jumlah": parse_num(m.group(4)),
            "harga_satuan": parse_num(m.group(5)),
            "total_harga": parse_num(m.group(6))
        })

    # --- 6️⃣ Validasi hasil parsing ---
    if not items:
        print("⚠️ Tidak ada item yang berhasil diparsing dari PDF.")
        return jsonify({"status": "error", "message": "Tidak ada item yang berhasil dibaca dari PDF."})

    # Urutkan A-Z untuk konsistensi tampilan
    items.sort(key=lambda x: x["nama_barang"].lower())

    print(f"✅ Berhasil membaca {len(items)} item dari PDF.")
    return jsonify({"status": "ok", "data": items})

@app.route("/export_penjualan_pdf/<nomor_faktur>")
def export_penjualan_pdf(nomor_faktur):
    import io
    from datetime import datetime
 
    # === Buat buffer dan style dasar dulu ===
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Header", fontSize=11, alignment=1, leading=13))
    styles.add(ParagraphStyle(name="NormalLeft", fontSize=10, alignment=0, leading=12))
    styles.add(ParagraphStyle(name="BoldCenter", fontSize=10, alignment=1, leading=12))

    # === Ambil setting layout (dari tabel LayoutSetting) ===
    conn2 = sqlite3.connect(DB_PATH)
    conn2.row_factory = sqlite3.Row
    cur2 = conn2.cursor()
    cur2.execute("SELECT * FROM LayoutSetting LIMIT 1")
    setting = cur2.fetchone()
    conn2.close()

    if setting:
        header_align = {"LEFT": 0, "CENTER": 1, "RIGHT": 2}.get(setting["posisi_header"], 1)
        font_size = setting["font_size_header"]
        leading = setting["jarak_header"]
        styles.add(ParagraphStyle(name="HeaderDynamic", fontSize=font_size, alignment=header_align, leading=leading))
    else:
        # fallback jika setting belum ada
        styles.add(ParagraphStyle(name="HeaderDynamic", fontSize=11, alignment=1, leading=13))

    # === Ambil data header dari Penjualan ===
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT id, tanggal_jual, konsumen, nomor_faktur_beli
        FROM Penjualan
        WHERE nomor_faktur_beli = ?
    """, (nomor_faktur,))
    penjualan = cur.fetchone()

    if not penjualan:
        conn.close()
        return f"❌ Faktur {nomor_faktur} tidak ditemukan di tabel Penjualan.", 404

    penjualan_id = penjualan["id"]

    # === Ambil item barang ===
    cur.execute("""
        SELECT nama_barang, satuan, jumlah, harga_jual
        FROM ItemPenjualan
        WHERE penjualan_id = ?
    """, (penjualan_id,))
    items = [dict(r) for r in cur.fetchall()]
    conn.close()

    if not items:
        return f"⚠️ Tidak ada item untuk faktur {nomor_faktur}.", 404

    # === Mulai isi konten PDF ===
    content = []

    # === Header koperasi (diambil dari LayoutSetting) ===
    if setting:
        koperasi_info = [
            setting["nama_koperasi"],
            setting["alamat"],
            setting["kontak"]
        ]
    else:
        koperasi_info = [
            "KOPERASI KONSUMEN MARSUKI RAGIL MANDIRI",
            "Jl Ryacudu gg Surya Alam no 09 Korpri Raya, Kec. Sukarame Kota Bandar Lampung",
            "081262014034"
        ]

    for line in koperasi_info:
        content.append(Paragraph(line, styles["HeaderDynamic"]))

    content.append(Spacer(1, 10))

    # === Info faktur ===
    content.append(Paragraph("<b>FAKTUR</b>", styles["BoldCenter"]))
    content.append(Paragraph(f"<b>#{nomor_faktur}</b>", styles["BoldCenter"]))
    content.append(Spacer(1, 8))
    content.append(Paragraph(f"Tanggal: {penjualan['tanggal_jual']}", styles["NormalLeft"]))
    content.append(Paragraph(f"Tagih Kepada: {penjualan['konsumen']}", styles["NormalLeft"]))
    content.append(Spacer(1, 10))

    # === Tabel item ===
    table_data = [["#", "Item", "Unit", "Kuantitas", "Biaya Satuan", "Total"]]
    subtotal = 0
    for i, item in enumerate(items, 1):
        total = item["jumlah"] * item["harga_jual"]
        subtotal += total
        table_data.append([
            str(i),
            item["nama_barang"],
            item["satuan"],
            f"{item['jumlah']:,.0f}",
            f"Rp{item['harga_jual']:,.0f}",
            f"Rp{total:,.0f}"
        ])

    table = Table(table_data, colWidths=[1*cm, 6*cm, 2*cm, 2.5*cm, 3*cm, 3*cm])
    table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("ALIGN", (1,1), (1,-1), "LEFT"),
        ("ALIGN", (4,1), (-1,-1), "RIGHT"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
    ]))
    content.append(table)
    content.append(Spacer(1, 10))

    # === Info pembayaran ===
    content.append(Paragraph("<b>Info Pembayaran</b>", styles["NormalLeft"]))
    content.append(Paragraph("Bri 579601000052568", styles["NormalLeft"]))
    content.append(Paragraph("An koperasi konsumen marsuki ragil mandiri", styles["NormalLeft"]))
    content.append(Paragraph("Bsi 7307554806", styles["NormalLeft"]))
    content.append(Paragraph("An/ koperasi marsuki ragil mandiri", styles["NormalLeft"]))
    content.append(Spacer(1, 8))

    # === Subtotal & Total ===
    total_str = f"Rp{subtotal:,.0f}"
    content.append(Paragraph(f"<b>Subtotal:</b> {total_str}", styles["NormalLeft"]))
    content.append(Paragraph(f"<b>Total:</b> {total_str}", styles["NormalLeft"]))
    content.append(Paragraph(f"<b>Saldo Terutang:</b> {total_str}", styles["NormalLeft"]))
    content.append(Spacer(1, 12))

    # === Footer ===
    content.append(Paragraph(
        "KOPERASI KONSUMEN MARSUKI RAGIL MANDIRI<br/>"
        "KELURAHAN KORPRI RAYA KECAMATAN SUKARAME KOTA BANDAR LAMPUNG",
        styles["BoldCenter"]
    ))
    content.append(Paragraph(datetime.now().strftime("%d/%m/%Y"), styles["BoldCenter"]))

    # === Build dokumen ===
    doc.build(content)
    buffer.seek(0)

    filename = f"{nomor_faktur}_{datetime.now().strftime('%d_%m_%Y')}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")

@app.route("/debug_pdf", methods=["POST"])
def debug_pdf():
    from PyPDF2 import PdfReader

    if "file" not in request.files:
        return jsonify({"status": "error", "message": "File tidak ditemukan"})

    file = request.files["file"]
    reader = PdfReader(file)
    text = ""
    for page in reader.pages:
        text += page.extract_text() + "\n"

    # Ambil hanya 50 baris pertama biar nggak kepanjangan
    lines = [ln for ln in text.splitlines() if ln.strip()]
    preview = "\n".join(lines[:50])

    return jsonify({"status": "ok", "preview": preview})

@app.route("/export_penjualan_excel_template/<nomor_faktur>")
def export_penjualan_excel_template(nomor_faktur):
    import os, io
    from datetime import datetime
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.cell.cell import MergedCell

    # === Helper aman untuk tulis ke merged cell ===
    def write_safe(ws, cell_ref, value, align=None):
        cell = ws[cell_ref]
        if isinstance(cell, MergedCell):
            for merged in ws.merged_cells.ranges:
                if cell_ref in merged:
                    top_left = merged.coord.split(":")[0]
                    ws[top_left].value = value
                    if align:
                        ws[top_left].alignment = align
                    return
        else:
            ws[cell_ref].value = value
            if align:
                ws[cell_ref].alignment = align

    # === Pastikan template Excel tersedia ===
    base_path = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_path, "Master Kops.xlsx")

    if not os.path.exists(template_path):
        return f"❌ Template Excel tidak ditemukan: {template_path}", 500

    # === Ambil data dari database ===
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("SELECT * FROM Penjualan WHERE nomor_faktur_beli = ?", (nomor_faktur,))
    penjualan = cur.fetchone()

    if not penjualan:
        conn.close()
        return f"❌ Faktur {nomor_faktur} tidak ditemukan.", 404

    cur.execute("""
        SELECT nama_barang, satuan, jumlah, harga_jual
        FROM ItemPenjualan
        WHERE penjualan_id = ?
    """, (penjualan["id"],))
    rows = cur.fetchall()
    conn.close()

    # === GROUPING (Gabungkan nama barang yang sama) ===
    grouped = {}

    for r in rows:
        key = (r["nama_barang"].strip().lower(), r["satuan"].strip().lower())

        if key not in grouped:
            grouped[key] = {
                "nama_barang": r["nama_barang"],
                "satuan": r["satuan"],
                "jumlah": float(r["jumlah"]),
                "harga_jual": float(r["harga_jual"])
            }
        else:
            # qty ditambah
            grouped[key]["jumlah"] += float(r["jumlah"])

            # jika harga jual beda, tetap ambil harga jual terakhir yang dimasukkan
            grouped[key]["harga_jual"] = float(r["harga_jual"])

    items = list(grouped.values())

    # === Load workbook ===
    wb = load_workbook(template_path)
    ws = wb.active

    # === Header ===
    write_safe(ws, "F6", penjualan["nomor_faktur_beli"])
    write_safe(ws, "F7", penjualan["konsumen"])
    write_safe(ws, "F8", penjualan["tanggal_jual"])

    TEMPLATE_ROW = 9
    jumlah_items = len(items)

    # Geser gambar
    shift = max(0, jumlah_items - 1)
    if shift > 0:
        for img in ws._images:
            img.anchor._from.row += shift

    # Insert rows
    if jumlah_items > 1:
        ws.insert_rows(TEMPLATE_ROW + 1, jumlah_items - 1)

    # === Isi item ===
    for i, item in enumerate(items):
        row = TEMPLATE_ROW + i

        write_safe(ws, f"B{row}", i + 1)
        write_safe(ws, f"C{row}", item["nama_barang"])
        write_safe(ws, f"D{row}", item["satuan"], Alignment(horizontal="center"))
        write_safe(ws, f"E{row}", item["jumlah"], Alignment(horizontal="center"))
        write_safe(ws, f"F{row}", item["harga_jual"], Alignment(horizontal="right"))
        write_safe(ws, f"G{row}", item["jumlah"] * item["harga_jual"], Alignment(horizontal="right"))

    # === Total ===
    TOTAL_ROW = TEMPLATE_ROW + jumlah_items + 1
    total_semua = sum(i["jumlah"] * i["harga_jual"] for i in items)

    write_safe(ws, f"F{TOTAL_ROW}", "TOTAL", Alignment(horizontal="right"))
    write_safe(ws, f"G{TOTAL_ROW}", total_semua, Alignment(horizontal="right"))

    # === Output ===
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{nomor_faktur}_Blanko_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# === SIMULASI HARGA JUAL (TIDAK SIMPAN DATABASE) ===
@app.route("/simulasi_harga_jual")
def simulasi_harga_jual():
    return render_template(
        "simulasi_harga_jual.html",
        today=date.today().strftime("%Y-%m-%d")
    )

# === API: HITUNG SIMULASI HARGA JUAL ===
@app.route("/api/simulasi_harga_jual")
def api_simulasi_harga_jual():
    nama = request.args.get("nama", "").strip()
    satuan = request.args.get("satuan", "").strip()
    qty = float(request.args.get("qty", 1) or 1)

    if not nama:
        return jsonify({"harga_beli": 0, "harga_jual": 0, "total": 0})

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # 1️⃣ ambil harga jual terakhir
    cur.execute("""
        SELECT harga_jual, harga_beli
        FROM ItemPenjualan
        WHERE LOWER(nama_barang)=LOWER(?) AND satuan=?
        ORDER BY id DESC LIMIT 1
    """, (nama, satuan))
    row = cur.fetchone()

    if row:
        harga_jual = row[0]
        harga_beli = row[1]
    else:
        # 2️⃣ fallback: harga beli terakhir dari pembelian
        cur.execute("""
            SELECT harga_satuan
            FROM ItemBelanja
            WHERE LOWER(nama_barang)=LOWER(?) AND satuan=?
            ORDER BY id DESC LIMIT 1
        """, (nama, satuan))
        hb = cur.fetchone()
        harga_beli = hb[0] if hb else 0
        harga_jual = round(harga_beli * 1.2) if harga_beli else 0

    conn.close()

    total = qty * harga_jual

    return jsonify({
        "harga_beli": harga_beli,
        "harga_jual": harga_jual,
        "total": total
    })

# ... (kode sisanya tetap sama) ...
if __name__ == "__main__":
    init_db()
    app.run(
        host="0.0.0.0",  # ⬅️ WAJIB agar bisa diakses intranet
        port=5000,
        debug=False     # ⬅️ disarankan untuk intranet
    )

