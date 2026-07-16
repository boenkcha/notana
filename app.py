from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
import sqlite3
from datetime import date
import pandas as pd
from io import BytesIO
import io
from flask import send_file
from PyPDF2 import PdfReader
import re
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

app = Flask(__name__)
app.secret_key = "secret-nota"

DB_PATH = "nota_belanja.db"

# === DB SETUP ===
def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS NotaBelanja (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tanggal TEXT NOT NULL,
                supplier TEXT NOT NULL,
                nomor_faktur TEXT UNIQUE NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS ItemBelanja (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nota_id INTEGER,
                nama_barang TEXT NOT NULL,
                jumlah REAL,
                satuan TEXT,
                harga_satuan REAL,
                total_harga REAL,
                FOREIGN KEY(nota_id) REFERENCES NotaBelanja(id)
            )
        """)

                # === Tambahan untuk Penjualan ===
        cur.execute("""
            CREATE TABLE IF NOT EXISTS Penjualan (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tanggal_jual TEXT NOT NULL,
                konsumen TEXT NOT NULL,
                nomor_faktur_beli TEXT NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS ItemPenjualan (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                penjualan_id INTEGER,
                nama_barang TEXT,
                jumlah REAL,
                satuan TEXT,
                harga_beli REAL,
                harga_jual REAL,
                total_jual REAL,
                FOREIGN KEY(penjualan_id) REFERENCES Penjualan(id)
            )
        """)

                # === Tabel Penjualan RAB (skema baru, tidak merusak Penjualan lama) ===
        cur.execute("""
            CREATE TABLE IF NOT EXISTS PenjualanRAB (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tanggal_jual TEXT NOT NULL,
                konsumen TEXT NOT NULL,
                nomor_faktur_beli TEXT NOT NULL,
                catatan TEXT,
                dibuat_pada TEXT DEFAULT (datetime('now','localtime'))
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS ItemPenjualanRAB (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                penjualan_rab_id INTEGER NOT NULL,
                -- sisi BELI (readonly, referensi)
                nama_barang_beli TEXT,
                jumlah_beli REAL,
                satuan_beli TEXT,
                harga_beli REAL,
                -- sisi JUAL (bisa diubah nama, qty, satuan, harga)
                nama_barang_jual TEXT,
                jumlah_jual REAL,
                satuan_jual TEXT,
                harga_jual REAL,
                total_jual REAL,
                FOREIGN KEY(penjualan_rab_id) REFERENCES PenjualanRAB(id)
            )
        """)

                # === Tabel Estimasi Bahan ===
        cur.execute("""
            CREATE TABLE IF NOT EXISTS EstimasiBahan (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tanggal TEXT NOT NULL,
                nama_menu TEXT NOT NULL,
                nomor_estimasi TEXT UNIQUE NOT NULL,
                dibuat_pada TEXT DEFAULT (datetime('now','localtime'))
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS ItemEstimasiBahan (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                estimasi_id INTEGER NOT NULL,
                nama_barang TEXT NOT NULL,
                jumlah REAL,
                satuan TEXT,
                harga_satuan REAL,
                total_harga REAL,
                FOREIGN KEY(estimasi_id) REFERENCES EstimasiBahan(id)
            )
        """)

                # === Tabel pengaturan layout faktur ===
        cur.execute("""
            CREATE TABLE IF NOT EXISTS LayoutSetting (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nama_koperasi TEXT,
                alamat TEXT,
                kontak TEXT,
                posisi_header TEXT,
                jarak_header REAL DEFAULT 12,
                font_size_header INTEGER DEFAULT 11
            )
        """)

        conn.commit()
                # Pastikan kolom tambahan untuk layout tahap 2 ada
        try:
            cur.execute("ALTER TABLE LayoutSetting ADD COLUMN table_header_color TEXT DEFAULT 'lightgrey'")
        except:
            pass
        try:
            cur.execute("ALTER TABLE LayoutSetting ADD COLUMN font_size_table INTEGER DEFAULT 9")
        except:
            pass
        try:
            cur.execute("ALTER TABLE LayoutSetting ADD COLUMN font_size_total INTEGER DEFAULT 10")
        except:
            pass
        try:
            cur.execute("ALTER TABLE LayoutSetting ADD COLUMN footer_align TEXT DEFAULT 'CENTER'")
        except:
            pass
        try:
            cur.execute("ALTER TABLE LayoutSetting ADD COLUMN spacing_after_table REAL DEFAULT 10")
        except:
            pass


       

# === HOME FORM ===
@app.route("/", methods=["GET"])
def index():
    return render_template("index.html", today=date.today().strftime("%Y-%m-%d"))

# === HANDLE SUBMIT ===
@app.route("/prepare", methods=["POST"])
def prepare():
    tanggal = request.form.get("tanggal")
    supplier = request.form.get("supplier")
    nomor_faktur = request.form.get("nomor_faktur")
    count = int(request.form.get("count", 0))

    if not tanggal or not supplier or not nomor_faktur:
        flash("⚠️ Mohon lengkapi Tanggal, Supplier, dan Nomor Faktur sebelum menyimpan!", "warning")
        return redirect(url_for("index"))

    items = []
    for i in range(count):
        nama = request.form.get(f"nama_{i}", "").strip()
        jumlah = request.form.get(f"jumlah_{i}", "").strip()
        satuan = request.form.get(f"satuan_{i}", "").strip()
        harga = request.form.get(f"harga_{i}", "").strip()

        if not nama:
            continue

        if not satuan:
            flash(f"⚠️ Baris ke-{i+1}: satuan belum diisi!", "danger")
            return redirect(url_for("index"))

        try:
            jumlah = float(jumlah or 0)
            harga = float(harga or 0)
        except ValueError:
            flash(f"⚠️ Baris ke-{i+1}: jumlah atau harga tidak valid!", "danger")
            return redirect(url_for("index"))

        if harga <= 0:
            flash(f"⚠️ Baris ke-{i+1}: harga harus lebih dari 0!", "danger")
            return redirect(url_for("index"))

        total = jumlah * harga
        items.append((nama, jumlah, satuan, harga, total))

    if not items:
        flash("⚠️ Tidak ada barang yang diinputkan!", "warning")
        return redirect(url_for("index"))

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    try:
        # 🧹 Hapus nota lama dengan nomor faktur yang sama (jika ada)
        cur.execute("SELECT id FROM NotaBelanja WHERE nomor_faktur = ?", (nomor_faktur,))
        nota = cur.fetchone()
        if nota:
            nota_id = nota[0]
            cur.execute("DELETE FROM ItemBelanja WHERE nota_id = ?", (nota_id,))
            cur.execute("DELETE FROM NotaBelanja WHERE id = ?", (nota_id,))

        # 🆕 Buat nota baru
        cur.execute("""
            INSERT INTO NotaBelanja (tanggal, supplier, nomor_faktur)
            VALUES (?, ?, ?)
        """, (tanggal, supplier, nomor_faktur))
        nota_id = cur.lastrowid

        # 💾 Simpan item baru
        for nama, jumlah, satuan, harga, total in items:
            cur.execute("""
                INSERT INTO ItemBelanja (nota_id, nama_barang, jumlah, satuan, harga_satuan, total_harga)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (nota_id, nama, jumlah, satuan, harga, total))

        conn.commit()
        flash(f"✅ Nota {nomor_faktur} berhasil disimpan ulang ({len(items)} item).", "success")

    except Exception as e:
        conn.rollback()
        flash(f"❌ Terjadi kesalahan saat menyimpan: {e}", "danger")

    finally:
        conn.close()

    return redirect(url_for("index"))



# === AUTOCOMPLETE BARANG ===
@app.route("/autocomplete", methods=["GET"])
def autocomplete():
    keyword = request.args.get("term", "").strip().lower()
    if not keyword:
        return jsonify(results=[])

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT nama_barang 
        FROM ItemBelanja 
        WHERE LOWER(nama_barang) LIKE ?
        ORDER BY nama_barang LIMIT 10
    """, (f"%{keyword}%",))
    results = [row["nama_barang"] for row in cur.fetchall()]
    conn.close()
    return jsonify(results=results)

@app.route("/layout_wizard", methods=["GET", "POST"])
def layout_wizard():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if request.method == "POST":
        nama = request.form["nama_koperasi"]
        alamat = request.form["alamat"]
        kontak = request.form["kontak"]
        posisi = request.form["posisi_header"]
        jarak = request.form["jarak_header"]
        font_size = request.form["font_size_header"]
        table_color = request.form.get("table_header_color", "lightgrey")
        font_table = request.form.get("font_size_table", 9)
        font_total = request.form.get("font_size_total", 10)
        footer_align = request.form.get("footer_align", "CENTER")
        spacing_after = request.form.get("spacing_after_table", 10)

        cur.execute("DELETE FROM LayoutSetting")
        cur.execute("""
            INSERT INTO LayoutSetting (
                nama_koperasi, alamat, kontak, posisi_header, jarak_header, font_size_header,
                table_header_color, font_size_table, font_size_total, footer_align, spacing_after_table
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (nama, alamat, kontak, posisi, jarak, font_size, table_color, font_table, font_total, footer_align, spacing_after))
        conn.commit()
        conn.close()


        flash("✅ Layout faktur berhasil disimpan!", "success")
        return redirect("/layout_wizard")

    # Ambil data setting terakhir
    cur.execute("SELECT * FROM LayoutSetting LIMIT 1")
    setting = cur.fetchone()
    conn.close()
    return render_template("layout_wizard.html", setting=setting)

# === REPORT ===
@app.route("/report")
def report():
    tanggal = request.args.get("tanggal", date.today().strftime("%Y-%m-%d"))
    nomor_faktur = request.args.get("nomor_faktur", "").strip()

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    query = """
        SELECT n.nomor_faktur, n.tanggal, i.nama_barang, i.jumlah, i.satuan, 
               i.harga_satuan, i.total_harga
        FROM NotaBelanja n
        JOIN ItemBelanja i ON n.id = i.nota_id
        WHERE 1=1
    """
    params = []

    if nomor_faktur:
        query += " AND n.nomor_faktur LIKE ?"
        params.append(f"%{nomor_faktur}%")
    else:
        query += " AND n.tanggal = ?"
        params.append(tanggal)

    query += " ORDER BY n.nomor_faktur ASC"

    cur.execute(query, params)
    rows = cur.fetchall()
    total = sum(r["total_harga"] for r in rows) if rows else 0
    conn.close()

    return render_template("report.html",
                           tanggal=tanggal,
                           rows=rows,
                           total=total,
                           nomor_faktur=nomor_faktur)

# === FORM PENJUALAN ===
@app.route("/penjualan", methods=["GET", "POST"])
def penjualan():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT nomor_faktur FROM NotaBelanja ORDER BY nomor_faktur DESC")
    nota_list = [r["nomor_faktur"] for r in cur.fetchall()]
    conn.close()

    if request.method == "POST":
        tanggal_jual = request.form.get("tanggal_jual")
        konsumen = request.form.get("konsumen")
        nomor_faktur_beli = request.form.get("nomor_faktur_beli")
        count = int(request.form.get("count", 0))

        if not (tanggal_jual and konsumen and nomor_faktur_beli):
            flash("⚠️ Lengkapi semua kolom terlebih dahulu!", "warning")
            return redirect(url_for("penjualan"))

        items = []
        for i in range(count):
            nama = request.form.get(f"nama_{i}")
            jumlah = request.form.get(f"jumlah_{i}")
            satuan = request.form.get(f"satuan_{i}")
            harga_beli = request.form.get(f"harga_beli_{i}")
            harga_jual = request.form.get(f"harga_jual_{i}")

            if not nama or not harga_jual:
                continue

            try:
                jumlah = float(jumlah)
                harga_beli = float(harga_beli)
                harga_jual = float(harga_jual)
            except ValueError:
                continue

            total_jual = jumlah * harga_jual
            items.append((nama, jumlah, satuan, harga_beli, harga_jual, total_jual))

        if not items:
            flash("⚠️ Tidak ada barang yang diinputkan!", "warning")
            return redirect(url_for("penjualan"))

        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO Penjualan (tanggal_jual, konsumen, nomor_faktur_beli)
            VALUES (?, ?, ?)
        """, (tanggal_jual, konsumen, nomor_faktur_beli))
        penjualan_id = cur.lastrowid

        for nama, jumlah, satuan, harga_beli, harga_jual, total_jual in items:
            cur.execute("""
                INSERT INTO ItemPenjualan (penjualan_id, nama_barang, jumlah, satuan, harga_beli, harga_jual, total_jual)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (penjualan_id, nama, jumlah, satuan, harga_beli, harga_jual, total_jual))

        conn.commit()
        conn.close()

        flash(f"✅ Penjualan ke {konsumen} berhasil disimpan.", "success")
        return redirect(url_for("penjualan"))

    return render_template("penjualan.html", nota_list=nota_list, today=date.today().strftime("%Y-%m-%d"))


# === API: Ambil Detail Nota Pembelian ===
@app.route("/get_items_by_nota")
def get_items_by_nota():
    nota = request.args.get("nota", "").strip()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT i.nama_barang, i.jumlah, i.satuan, i.harga_satuan
        FROM NotaBelanja n
        JOIN ItemBelanja i ON n.id = i.nota_id
        WHERE n.nomor_faktur = ?
    """, (nota,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify(rows)


# === SIMPAN PENJUALAN ===
@app.route("/simpan_penjualan", methods=["POST"])
def simpan_penjualan():
    tanggal = request.form.get("tanggal")
    konsumen = request.form.get("konsumen")
    nota_beli = request.form.get("nota_beli")

    if not tanggal or not konsumen or not nota_beli:
        flash("⚠️ Lengkapi semua kolom terlebih dahulu!", "warning")
        return redirect("/penjualan")

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # Pastikan tabel ada
    cur.execute("""
        CREATE TABLE IF NOT EXISTS Penjualan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tanggal_jual TEXT NOT NULL,
            konsumen TEXT NOT NULL,
            nomor_faktur_beli TEXT UNIQUE NOT NULL
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS ItemPenjualan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            penjualan_id INTEGER,
            nama_barang TEXT,
            jumlah REAL,
            satuan TEXT,
            harga_beli REAL,
            harga_jual REAL,
            total_jual REAL,
            FOREIGN KEY(penjualan_id) REFERENCES Penjualan(id)
        )
    """)
    conn.commit()

    # Cek apakah faktur sudah ada
    cur.execute("SELECT id FROM Penjualan WHERE nomor_faktur_beli = ?", (nota_beli,))
    row = cur.fetchone()

    if row:
        penjualan_id = row[0]
        # Hapus semua item lama untuk faktur ini
        cur.execute("DELETE FROM ItemPenjualan WHERE penjualan_id = ?", (penjualan_id,))
        mode = "update"
    else:
        # Faktur baru
        cur.execute("""
            INSERT INTO Penjualan (tanggal_jual, konsumen, nomor_faktur_beli)
            VALUES (?, ?, ?)
        """, (tanggal, konsumen, nota_beli))
        penjualan_id = cur.lastrowid
        mode = "new"

    # Simpan item baru
    count = 0
    for key in request.form:
        if key.startswith("nama_"):
            index = key.split("_")[1]
            nama = request.form.get(f"nama_{index}", "").strip()
            jumlah = float(request.form.get(f"jumlah_{index}", 0) or 0)
            satuan = request.form.get(f"satuan_{index}", "").strip()
            harga_beli = float(request.form.get(f"harga_beli_{index}", 0) or 0)
            harga_jual = float(request.form.get(f"harga_jual_{index}", 0) or 0)
            total = jumlah * harga_jual

            if not nama or harga_jual <= 0:
                continue

            cur.execute("""
                INSERT INTO ItemPenjualan
                (penjualan_id, nama_barang, jumlah, satuan, harga_beli, harga_jual, total_jual)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (penjualan_id, nama, jumlah, satuan, harga_beli, harga_jual, total))
            count += 1

    conn.commit()
    conn.close()

    # Pesan flash
    if mode == "update":
        flash(f"✅ Faktur {nota_beli} diperbarui ({count} item disimpan, data lama dihapus).", "success")
    else:
        flash(f"✅ Penjualan baru disimpan ({count} item) untuk faktur {nota_beli}.", "success")

    return redirect("/penjualan")


@app.route("/cek_faktur_penjualan")
def cek_faktur_penjualan():
    nomor_faktur = request.args.get("nomor_faktur")
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT id FROM Penjualan WHERE nomor_faktur_beli = ?", (nomor_faktur,))
    ada = cur.fetchone()
    conn.close()
    return jsonify({"exists": bool(ada)})

# === API: Gabungkan ItemBelanja + ItemPenjualan berdasarkan Nomor Faktur ===
@app.route('/cari_faktur_penjualan')
def cari_faktur_penjualan():
    nomor_faktur = request.args.get('nomor_faktur')
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Ambil data nota pembelian
    cur.execute("""
        SELECT tanggal, supplier AS konsumen
        FROM NotaBelanja
        WHERE nomor_faktur = ?
    """, (nomor_faktur,))
    nota = cur.fetchone()

    if not nota:
        return jsonify({"status": "not_found"})

    # Ambil data item pembelian
    cur.execute("""
        SELECT nama_barang, jumlah, satuan, harga_satuan AS harga_beli
        FROM ItemBelanja
        WHERE nota_id = (SELECT id FROM NotaBelanja WHERE nomor_faktur = ?)
    """, (nomor_faktur,))
    items = [dict(row) for row in cur.fetchall()]

    # Cek apakah sudah pernah dijual (dari tabel ItemPenjualanRAB, bukan ItemPenjualan)
    for item in items:
        cur.execute("""
            SELECT i.harga_jual
            FROM ItemPenjualanRAB i
            WHERE LOWER(i.nama_barang_jual) = LOWER(?)
            ORDER BY i.id DESC LIMIT 1
        """, (item['nama_barang'],))
        jual = cur.fetchone()
        item['harga_jual'] = jual['harga_jual'] if jual else None

    conn.close()
    return jsonify({
        "status": "ok",
        "tanggal": nota["tanggal"],
        "konsumen": nota["konsumen"],
        "items": items
    })

@app.route("/suggest_harga", methods=["POST"])
def suggest_harga():
    data = request.get_json()
    items = data.get("items", [])

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    hasil = []
    for item in items:
        nama = item.get("nama")
        satuan = item.get("satuan")

        # Ambil harga jual terakhir
        cur.execute("""
            SELECT harga_jual FROM ItemPenjualan
            WHERE nama_barang = ? COLLATE NOCASE AND satuan = ? COLLATE NOCASE
            ORDER BY id DESC LIMIT 1
        """, (nama, satuan))

        row = cur.fetchone()

        if row and row[0]:
            hasil.append(row[0])
        else:
            # fallback: ambil harga beli + 20%
            cur.execute("""
                SELECT harga_beli FROM ItemPenjualan
                WHERE nama_barang = ? COLLATE NOCASE AND satuan = ? COLLATE NOCASE
                ORDER BY id DESC LIMIT 1
            """, (nama, satuan))

            hb = cur.fetchone()
            hasil.append(round(hb[0] * 1.2) if hb else None)

    conn.close()
    return jsonify({"status": "ok", "saran": hasil})


# === API: Suggest Harga Jual RAB (dari tabel ItemPenjualanRAB) ===
@app.route("/suggest_harga_rab", methods=["POST"])
def suggest_harga_rab():
    """
    Urutan prioritas per item:
    1. Harga jual dari ItemPenjualanRAB dengan nota yang SAMA (nomor_faktur_beli sama)
    2. Harga jual terakhir dari ItemPenjualanRAB berdasarkan nama barang (cross-nota)
    3. 0 (belum ada riwayat sama sekali)
    Tidak lagi mengambil dari tabel ItemPenjualan.
    """
    data = request.get_json()
    items        = data.get("items", [])
    nomor_faktur = data.get("nomor_faktur", "").strip()

    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()

    hasil = []
    for item in items:
        nama   = (item.get("nama")   or "").strip()
        satuan = (item.get("satuan") or "").strip()

        harga = None

        # ── Prioritas 1: harga dari nota yang SAMA ──
        if nomor_faktur:
            cur.execute("""
                SELECT i.harga_jual
                FROM ItemPenjualanRAB i
                JOIN PenjualanRAB p ON i.penjualan_rab_id = p.id
                WHERE p.nomor_faktur_beli = ?
                  AND LOWER(i.nama_barang_jual) = LOWER(?)
                ORDER BY i.id DESC
                LIMIT 1
            """, (nomor_faktur, nama))
            row = cur.fetchone()
            if row and row[0]:
                harga = row[0]

        # ── Prioritas 2: harga terakhir cross-nota dari ItemPenjualanRAB ──
        if harga is None:
            cur.execute("""
                SELECT i.harga_jual
                FROM ItemPenjualanRAB i
                WHERE LOWER(i.nama_barang_jual) = LOWER(?)
                ORDER BY i.id DESC
                LIMIT 1
            """, (nama,))
            row = cur.fetchone()
            if row and row[0]:
                harga = row[0]

        # ── Prioritas 3: 0 ──
        hasil.append(harga if harga is not None else 0)

    conn.close()
    return jsonify({"status": "ok", "saran": hasil})


@app.route("/export_penjualan_excel")
def export_penjualan_excel():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Ambil data gabungan Penjualan + ItemPenjualan
    query = """
        SELECT 
            p.tanggal_jual AS Tanggal,
            p.nomor_faktur_beli AS Nomor_Faktur,
            p.konsumen AS Konsumen,
            i.nama_barang AS Nama_Barang,
            i.jumlah AS Jumlah,
            i.satuan AS Satuan,
            i.harga_beli AS Harga_Beli,
            i.harga_jual AS Harga_Jual,
            i.total_jual AS Total_Jual
        FROM Penjualan p
        JOIN ItemPenjualan i ON p.id = i.penjualan_id
        ORDER BY p.tanggal_jual DESC, p.nomor_faktur_beli ASC
    """
    cur.execute(query)
    rows = cur.fetchall()
    conn.close()

    if not rows:
        flash("⚠️ Tidak ada data penjualan untuk diekspor.", "warning")
        return redirect("/penjualan")

    # Buat dataframe dengan urutan kolom yang rapi
    df = pd.DataFrame(rows, columns=[
        "Tanggal", "Nomor_Faktur", "Konsumen",
        "Nama_Barang", "Jumlah", "Satuan",
        "Harga_Beli", "Harga_Jual", "Total_Jual"
    ])

    # Format angka
    df["Jumlah"] = df["Jumlah"].astype(float)
    df["Harga_Beli"] = df["Harga_Beli"].astype(float)
    df["Harga_Jual"] = df["Harga_Jual"].astype(float)
    df["Total_Jual"] = df["Total_Jual"].astype(float)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="Data Penjualan", index=False)
        workbook = writer.book
        worksheet = writer.sheets["Data Penjualan"]

        # Format Header
        header_fmt = workbook.add_format({
            "bold": True,
            "align": "center",
            "valign": "vcenter",
            "fg_color": "#B7DEE8",
            "border": 1
        })

        # Format isi angka
        money_fmt = workbook.add_format({
            "num_format": '#,##0',
            "align": "right",
            "border": 1
        })

        # Format isi teks
        text_fmt = workbook.add_format({
            "align": "left",
            "border": 1
        })

        # Terapkan format header dan lebar kolom
        for col_num, col_name in enumerate(df.columns):
            worksheet.write(0, col_num, col_name, header_fmt)

            if col_name in ["Jumlah", "Harga_Beli", "Harga_Jual", "Total_Jual"]:
                worksheet.set_column(col_num, col_num, 15, money_fmt)
            elif col_name in ["Tanggal", "Nomor_Faktur"]:
                worksheet.set_column(col_num, col_num, 18, text_fmt)
            else:
                worksheet.set_column(col_num, col_num, 20, text_fmt)

        # Tambahkan autofilter di header
        worksheet.autofilter(0, 0, len(df), len(df.columns) - 1)

        # Tambahkan total di bawah tabel
        total_row = len(df) + 1
        worksheet.write(total_row, 7, "TOTAL :", header_fmt)
        worksheet.write_formula(total_row, 8, f"=SUM(I2:I{total_row})", money_fmt)

    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="Data_Penjualan_Rapi.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/export_excel")
def export_excel():
    tanggal = request.args.get("tanggal", date.today().strftime("%Y-%m-%d"))
    nomor_faktur = request.args.get("nomor_faktur", "").strip()

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    query = """
        SELECT n.nomor_faktur, n.tanggal, i.nama_barang, i.jumlah, i.satuan,
               i.harga_satuan, i.total_harga
        FROM NotaBelanja n
        JOIN ItemBelanja i ON n.id = i.nota_id
        WHERE 1=1
    """
    params = []

    if nomor_faktur:
        query += " AND n.nomor_faktur LIKE ?"
        params.append(f"%{nomor_faktur}%")
    else:
        query += " AND n.tanggal = ?"
        params.append(tanggal)

    query += " ORDER BY n.nomor_faktur ASC"

    cur.execute(query, params)
    rows = cur.fetchall()
    conn.close()

    if not rows:
        flash("Tidak ada data untuk diexport", "warning")
        return redirect("/report")

    df = pd.DataFrame(rows)
    path = "report_harian.xlsx"
    df.to_excel(path, index=False)

    return send_file(path, as_attachment=True)

def get_most_frequent_price(nama_barang):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        SELECT harga_satuan, COUNT(*) AS freq
        FROM ItemBelanja
        WHERE LOWER(nama_barang) = LOWER(?)
        GROUP BY harga_satuan
        ORDER BY freq DESC
        LIMIT 1
    """, (nama_barang,))
    result = cur.fetchone()
    conn.close()
    return result[0] if result else None

@app.route('/cari_barang')
def cari_barang():
    from rapidfuzz import process, fuzz
    q    = request.args.get('q',    '').strip().lower()
    mode = request.args.get('mode', 'beli')   # 'beli' (default) atau 'jual'

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if mode == 'jual':
        # Ambil dari sisi jual RAB — harga jual terakhir per nama
        cur.execute("""
            SELECT id,
                   nama_barang_jual AS nama_barang,
                   satuan_jual      AS satuan,
                   harga_jual       AS harga_satuan
            FROM ItemPenjualanRAB
            WHERE nama_barang_jual IS NOT NULL AND nama_barang_jual <> ''
            ORDER BY id DESC
        """)
    else:
        # Default: ambil dari sisi beli — harga beli terakhir per nama
        cur.execute("""
            SELECT id, nama_barang, satuan, harga_satuan
            FROM ItemBelanja
            WHERE nama_barang IS NOT NULL AND nama_barang <> ''
            ORDER BY id DESC
        """)
    rows = cur.fetchall()
    conn.close()

    if not q:
        return jsonify(results=[])

    q_words = q.lower().split()
    all_names = [d['nama_barang'] for d in rows]

    # 1. Substring match: semua kata query harus ada di nama (prioritas utama)
    substring_set = {
        r["nama_barang"] for r in rows
        if all(w in r["nama_barang"].lower() for w in q_words)
    }

    # 2. Fuzzy match: untuk toleransi typo (dari seluruh data)
    fuzzy_hasil = process.extract(q, all_names, limit=50, scorer=fuzz.token_set_ratio)
    fuzzy_set = {nama for nama, skor, *_ in fuzzy_hasil if skor >= 55}

    # Gabungkan: substring dulu, lalu fuzzy tambahan
    hasil_filtered = list(dict.fromkeys(
        [n for n in all_names if n in substring_set] +
        [n for n in all_names if n in fuzzy_set and n not in substring_set]
    ))

    # === UNIQUE BERDASARKAN RECORD TERBARU ===
    unique_latest = {}
    for r in rows:
        nama = r["nama_barang"]
        if nama in hasil_filtered and nama not in unique_latest:
            unique_latest[nama] = r

    # Urutkan sesuai hasil_filtered (substring duluan)
    results = []
    for nama in hasil_filtered:
        if nama in unique_latest:
            r = unique_latest[nama]
            results.append({
                "nama": r["nama_barang"],
                "satuan": r["satuan"],
                "harga": r["harga_satuan"]
            })
        if len(results) >= 8:
            break

    return jsonify(results=results)

@app.route("/get_harga_terbanyak")
def get_harga_terbanyak():
    nama = request.args.get("nama", "").strip()
    if not nama:
        return jsonify({"harga": None})
    harga = get_most_frequent_price(nama)
    return jsonify({"harga": harga})

@app.route("/get_harga_terbaru")
def get_harga_terbaru():
    nama = request.args.get("nama", "").strip()
    if not nama:
        return jsonify({"harga": None})

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # Join ke NotaBelanja agar pengurutan berdasarkan tanggal faktur yang benar,
    # bukan sekadar id (id bisa tidak urut jika nota lama di-edit ulang)
    cur.execute("""
        SELECT ib.harga_satuan
        FROM ItemBelanja ib
        JOIN NotaBelanja nb ON nb.id = ib.nota_id
        WHERE ib.nama_barang = ?
        ORDER BY nb.tanggal DESC, ib.id DESC
        LIMIT 1
    """, (nama,))

    row = cur.fetchone()
    conn.close()

    return jsonify({
        "harga": row[0] if row else None
    })

@app.route('/cek_fuzzy_barang')
def cek_fuzzy_barang():
    try:
        from fuzzywuzzy import fuzz, process
    except ImportError:
        from rapidfuzz import fuzz, process  # fallback jika fuzzywuzzy tidak ada

    q = request.args.get('q', '').strip()
    if not q:
        return jsonify(match=None, score=0)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT nama_barang FROM ItemBelanja WHERE nama_barang IS NOT NULL AND nama_barang <> ''")
    all_names = [row['nama_barang'] for row in cur.fetchall()]
    conn.close()

    if not all_names:
        return jsonify(match=None, score=0)

    match, score, *_ = process.extractOne(q, all_names, scorer=fuzz.token_set_ratio)
    if score >= 85:
        return jsonify(match=match, score=score)
    else:
        return jsonify(match=None, score=score)

@app.route('/saran_barang')
def saran_barang():
    try:
        from fuzzywuzzy import fuzz, process
    except ImportError:
        from rapidfuzz import fuzz, process

    q = request.args.get('q', '').strip()
    if not q:
        return jsonify(suggestions=[])

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT nama_barang FROM ItemBelanja WHERE nama_barang IS NOT NULL AND nama_barang <> ''")
    all_names = [row['nama_barang'] for row in cur.fetchall()]
    conn.close()

    if not all_names:
        return jsonify(suggestions=[])

    hasil = process.extract(q, all_names, limit=5, scorer=fuzz.token_set_ratio)
    hasil_filtered = [h for h in hasil if h[1] >= 70]

    suggestions = [{"nama": h[0], "score": h[1]} for h in hasil_filtered]
    return jsonify(suggestions=suggestions)

@app.route("/export_penjualan_excel_by_faktur")
def export_penjualan_excel_by_faktur():
    import io
    import pandas as pd
    from datetime import datetime

    nomor_faktur = request.args.get("nomor_faktur", "").strip()
    if not nomor_faktur:
        flash("⚠️ Nomor faktur wajib diisi!", "warning")
        return redirect("/penjualan")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # === Header Faktur ===
    cur.execute("SELECT id, konsumen, tanggal_jual FROM Penjualan WHERE nomor_faktur_beli = ?", (nomor_faktur,))
    penjualan = cur.fetchone()
    if not penjualan:
        conn.close()
        flash(f"⚠️ Faktur {nomor_faktur} tidak ditemukan.", "warning")
        return redirect("/penjualan")

    # === Data Item ===
    cur.execute("""
        SELECT 
            nama_barang AS 'Nama Barang',
            jumlah AS 'Jumlah',
            satuan AS 'Satuan',
            harga_beli AS 'Harga Beli',
            (jumlah * harga_beli) AS 'Total Beli',
            harga_jual AS 'Harga Jual',
            total_jual AS 'Subtotal'
        FROM ItemPenjualan
        WHERE penjualan_id = ?
    """, (penjualan["id"],))
    rows = cur.fetchall()
    conn.close()

    if not rows:
        flash("⚠️ Tidak ada item untuk faktur ini.", "warning")
        return redirect("/penjualan")

    df = pd.DataFrame(rows)

    # === Buat file Excel ===
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        workbook = writer.book
        worksheet = workbook.add_worksheet("Penjualan")
        writer.sheets["Penjualan"] = worksheet

        # === FORMAT ===
        title_fmt = workbook.add_format({"bold": True, "font_size": 14, "align": "center"})
        header_fmt = workbook.add_format({
            "bold": True, "align": "center", "valign": "vcenter",
            "bg_color": "#cfe2f3", "border": 1
        })
        text_fmt = workbook.add_format({"border": 1})
        money_fmt = workbook.add_format({"num_format": '#,##0', "align": "right", "border": 1})
        info_fmt = workbook.add_format({"bold": True, "align": "left"})
        total_label_fmt = workbook.add_format({
            "bold": True, "align": "right", "border": 1, "bg_color": "#fce5cd"
        })

        # === JUDUL ===
        worksheet.merge_range("A1:G1", "LAPORAN PENJUALAN PER FAKTUR", title_fmt)

        # === INFO FAKTUR ===
        worksheet.write("A3", "Nomor Faktur:", info_fmt)
        worksheet.write("B3", nomor_faktur)
        worksheet.write("A4", "Tanggal Jual:", info_fmt)
        worksheet.write("B4", penjualan["tanggal_jual"])
        worksheet.write("A5", "Konsumen:", info_fmt)
        worksheet.write("B5", penjualan["konsumen"])

        # === HEADER KOLOM ===
        for col_num, col_name in enumerate(df.columns):
            worksheet.write(6, col_num, col_name, header_fmt)

        # === ISI DATA ===
        for row_num, row_data in enumerate(df.values):
            for col_num, cell_value in enumerate(row_data):
                if df.columns[col_num] in ["Harga Beli", "Total Beli", "Harga Jual", "Subtotal"]:
                    worksheet.write(row_num + 7, col_num, cell_value, money_fmt)
                else:
                    worksheet.write(row_num + 7, col_num, cell_value, text_fmt)

        # === FORMAT KOLOM ===
        worksheet.set_column("A:A", 25)
        worksheet.set_column("B:B", 10)
        worksheet.set_column("C:C", 10)
        worksheet.set_column("D:D", 15)
        worksheet.set_column("E:E", 15)
        worksheet.set_column("F:F", 15)
        worksheet.set_column("G:G", 15)

        # === TOTAL DAN LABA ===
        total_row = len(df) + 7
        worksheet.write(total_row, 3, "TOTAL BELI :", total_label_fmt)
        worksheet.write_formula(total_row, 4, f"=SUM(E8:E{total_row})", money_fmt)

        worksheet.write(total_row + 1, 3, "TOTAL PENJUALAN :", total_label_fmt)
        worksheet.write_formula(total_row + 1, 6, f"=SUM(G8:G{total_row})", money_fmt)

        worksheet.write(total_row + 2, 3, "LABA KOTOR :", total_label_fmt)
        worksheet.write_formula(total_row + 2, 6, f"=G{total_row+2}-E{total_row+1}", money_fmt)

    output.seek(0)
    filename = f"Penjualan_{nomor_faktur}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# === API: Ambil Nota Lengkap (header + item) ===
@app.route("/get_nota_detail")
def get_nota_detail():
    nomor_faktur = request.args.get("nomor_faktur", "").strip()
    if not nomor_faktur:
        return jsonify({"status": "error", "message": "Nomor faktur kosong"})

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT id, tanggal, supplier, nomor_faktur
        FROM NotaBelanja
        WHERE nomor_faktur = ?
    """, (nomor_faktur,))
    nota = cur.fetchone()

    if not nota:
        conn.close()
        return jsonify({"status": "not_found", "message": f"Nota {nomor_faktur} tidak ditemukan"})

    cur.execute("""
        SELECT nama_barang, jumlah, satuan, harga_satuan, total_harga
        FROM ItemBelanja
        WHERE nota_id = ?
        ORDER BY id ASC
    """, (nota["id"],))
    items = [dict(r) for r in cur.fetchall()]
    conn.close()

    return jsonify({
        "status": "ok",
        "tanggal": nota["tanggal"],
        "supplier": nota["supplier"],
        "nomor_faktur": nota["nomor_faktur"],
        "items": items
    })

@app.route("/get_all_nota")
def get_all_nota():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT nomor_faktur, supplier
        FROM NotaBelanja
        ORDER BY tanggal DESC, id DESC
        LIMIT 500
    """)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"status": "ok", "notas": rows})

# === API: Riwayat Harga Barang (gabungan pembelian + penjualan) ===
@app.route("/riwayat_harga")
def riwayat_harga():
    nama = request.args.get("nama", "").strip()
    if not nama:
        return jsonify({"status": "error", "message": "Nama barang kosong"})

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # === BELI: 1 baris per tanggal ===
    cur.execute("""
        SELECT n.tanggal AS tanggal, n.supplier AS pihak,
               i.harga_satuan AS harga_beli, NULL AS harga_jual, 'BELI' AS jenis
        FROM ItemBelanja i
        JOIN NotaBelanja n ON i.nota_id = n.id
        WHERE LOWER(i.nama_barang) = LOWER(?)
        GROUP BY n.tanggal
        ORDER BY n.tanggal DESC
    """, (nama,))
    beli = [dict(r) for r in cur.fetchall()]

    # === JUAL (Penjualan biasa): 1 baris per tanggal ===
    cur.execute("""
        SELECT p.tanggal_jual AS tanggal, p.konsumen AS pihak,
               NULL AS harga_beli, i.harga_jual AS harga_jual, 'JUAL' AS jenis
        FROM ItemPenjualan i
        JOIN Penjualan p ON i.penjualan_id = p.id
        WHERE LOWER(i.nama_barang) = LOWER(?)
        GROUP BY p.tanggal_jual
        ORDER BY p.tanggal_jual DESC
    """, (nama,))
    jual = [dict(r) for r in cur.fetchall()]

    # === JUAL RAB (PenjualanRAB): 1 baris per tanggal ===
    cur.execute("""
        SELECT p.tanggal_jual AS tanggal, p.konsumen AS pihak,
               i.harga_beli AS harga_beli, i.harga_jual AS harga_jual, 'JUAL_RAB' AS jenis
        FROM ItemPenjualanRAB i
        JOIN PenjualanRAB p ON i.penjualan_rab_id = p.id
        WHERE LOWER(i.nama_barang_jual) = LOWER(?)
        GROUP BY p.tanggal_jual
        ORDER BY p.tanggal_jual DESC
    """, (nama,))
    jual_rab = [dict(r) for r in cur.fetchall()]

    data = beli + jual + jual_rab
    data.sort(key=lambda x: x["tanggal"] or "", reverse=True)

    conn.close()
    return jsonify({"status": "ok", "data": data})


@app.route("/cari_item")
def cari_item_page():
    return render_template("cari_item.html")

@app.route("/simpan_arsip_pdf", methods=["POST"])
def simpan_arsip_pdf():
    try:
        file = request.files.get("file")
        nomor_faktur = request.form.get("nomor_faktur", "").strip()

        if not file:
            return jsonify({"status": "error", "message": "File PDF tidak ditemukan di request."})
        if not nomor_faktur:
            return jsonify({"status": "error", "message": "Nomor faktur kosong."})

        # Pastikan folder arsip ada
        folder_arsip = os.path.join(app.static_folder, "arsip_penjualan")
        os.makedirs(folder_arsip, exist_ok=True)

        # Buat nama file aman (ganti / jadi _)
        safe_name = f"{nomor_faktur.replace('/', '_')}.pdf"
        file_path = os.path.join(folder_arsip, safe_name)

        # Simpan file PDF
        file.save(file_path)
        print(f"✅ Arsip tersimpan: {file_path}")

        return jsonify({"status": "ok", "message": "Arsip berhasil disimpan.", "file_url": f"/static/arsip_penjualan/{safe_name}"})
    except Exception as e:
        print("❌ ERROR simpan_arsip_pdf:", str(e))
        return jsonify({"status": "error", "message": str(e)})

    

import os
from flask import send_from_directory

@app.route("/get_daftar_arsip", methods=["GET"])
def get_daftar_arsip():
    folder = os.path.join(app.root_path, "static", "arsip_penjualan")
    if not os.path.exists(folder):
        return jsonify({"status": "ok", "data": []})

    files = []
    for f in sorted(os.listdir(folder)):
        if f.lower().endswith(".pdf"):
            files.append({
                "nama": f,
                "url": url_for('static', filename=f"arsip_penjualan/{f}")
            })

    return jsonify({"status": "ok", "data": files})


@app.route("/cek_arsip_pdf")
def cek_arsip_pdf():
    nomor_faktur = request.args.get("nomor_faktur", "").strip()
    if not nomor_faktur:
        return jsonify({"status": "error", "message": "Nomor faktur kosong"})

    folder_arsip = os.path.join(app.static_folder, "arsip_penjualan")
    safe_name = f"{nomor_faktur.replace('/', '_')}.pdf"
    file_path = os.path.join(folder_arsip, safe_name)

    if os.path.exists(file_path):
        file_url = f"/static/arsip_penjualan/{safe_name}"
        print("✅ File ditemukan:", file_url)  # Tambahkan ini untuk lihat hasil di console Flask
        return jsonify({"status": "ok", "file_url": file_url})
    else:
        print("❌ Tidak ditemukan:", file_path)
        return jsonify({"status": "not_found"})


@app.route("/baca_pdf_arsip")
def baca_pdf_arsip():
    nomor_faktur = request.args.get("nomor_faktur", "").strip()
    if not nomor_faktur:
        return jsonify({"status": "error", "message": "Nomor faktur kosong"})

    folder_arsip = os.path.join("static", "arsip_penjualan")
    safe_name = f"{nomor_faktur.replace('/', '_')}.pdf"
    file_path = os.path.join(folder_arsip, safe_name)

    if os.path.exists(file_path):
        return jsonify({
            "status": "ok",
            "file_url": f"/{file_path.replace(os.sep, '/')}"
        })
    else:
        return jsonify({"status": "not_found"})


@app.route("/get_arsip_pdf")
def get_arsip_pdf():
    nomor_faktur = request.args.get("nomor_faktur", "").strip()
    if not nomor_faktur:
        return jsonify({"status": "error", "message": "Nomor faktur kosong"})

    pdf_path = os.path.join("arsip", f"{nomor_faktur}.pdf")
    if os.path.exists(pdf_path):
        return jsonify({"status": "ok", "path": "/" + pdf_path.replace("\\", "/")})
    else:
        return jsonify({"status": "not_found"})


@app.route("/hapus_arsip_pdf", methods=["POST"])
def hapus_arsip_pdf():
    nama = request.form.get("nama", "")
    if not nama:
        return jsonify({"status": "error", "message": "Nama file tidak diberikan."})

    arsip_folder = os.path.join("static", "arsip_pdf")
    file_path = os.path.join(arsip_folder, nama)

    if not os.path.exists(file_path):
        return jsonify({"status": "error", "message": "File tidak ditemukan."})

    try:
        os.remove(file_path)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


@app.route("/get_pdf_arsip")
def get_pdf_arsip():
    nomor_faktur = request.args.get("nomor_faktur", "").strip()
    if not nomor_faktur:
        return jsonify({"status": "error", "message": "Nomor faktur kosong"})

    pdf_path = f"static/arsip/{nomor_faktur}.pdf"
    if not os.path.exists(pdf_path):
        return jsonify({"status": "not_found", "message": "Arsip PDF tidak ditemukan."})

    # Baca teks dari PDF (optional)
    try:
        reader = PdfReader(pdf_path)
        pdf_text = ""
        for page in reader.pages:
            pdf_text += page.extract_text() + "\n"
    except Exception as e:
        pdf_text = f"Gagal membaca isi PDF: {e}"

    return jsonify({
        "status": "ok",
        "pdf_url": f"/static/arsip/{nomor_faktur}.pdf",
        "pdf_text": pdf_text
    })

@app.route("/get_item_penjualan")
def get_item_penjualan():
    try:
        nomor_faktur = request.args.get("nomor_faktur", "").strip()
        if not nomor_faktur:
            return jsonify({
                "status": "error",
                "message": "Nomor faktur kosong."
            })

        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        cur.execute("""
            SELECT 
                i.nama_barang,
                i.jumlah,
                i.satuan,
                i.harga_jual,
                i.total_jual
            FROM ItemPenjualan i
            JOIN Penjualan p ON i.penjualan_id = p.id
            WHERE p.nomor_faktur_beli = ?
        """, (nomor_faktur,))
        rows = cur.fetchall()
        conn.close()

        if not rows:
            return jsonify({
                "status": "error",
                "message": f"Tidak ditemukan item untuk nota penjualan dengan nomor faktur {nomor_faktur}."
            })

        data = [dict(row) for row in rows]
        return jsonify({"status": "ok", "data": data})

    except sqlite3.OperationalError as e:
        return jsonify({
            "status": "error",
            "message": f"Kesalahan database: {str(e)}"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Terjadi kesalahan: {str(e)}"
        })


@app.route("/upload_pdf", methods=["POST"])
def upload_pdf():
    from PyPDF2 import PdfReader
    import re, os, tempfile, traceback

    print("=== DEBUG upload_pdf ===")

    # --- 1️⃣ Pastikan file ada ---
    if "file" not in request.files:
        print("🚫 Tidak ada file PDF dalam request.")
        return jsonify({"status": "error", "message": "Tidak ada file PDF yang dikirim."})

    file = request.files["file"]
    nomor_faktur = request.form.get("nomor_faktur", "").strip() or "(tidak diketahui)"
    print(f"📦 Menerima file: {file.filename} | tipe: {file.content_type} | faktur: {nomor_faktur}")

    # --- 2️⃣ Simpan dulu ke file sementara agar bisa dibaca PyPDF2 ---
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name

        if os.path.getsize(tmp_path) == 0:
            print("🚫 File PDF kosong!")
            return jsonify({"status": "error", "message": "File PDF kosong atau gagal diambil dari browser."})
    except Exception as e:
        print("❌ Gagal menyimpan file sementara:", e)
        return jsonify({"status": "error", "message": "Gagal menyimpan file PDF sementara."})

    # --- 3️⃣ Baca isi PDF ---
    try:
        reader = PdfReader(tmp_path)
        print(f"📖 Jumlah halaman PDF: {len(reader.pages)}")

        text = ""
        for page in reader.pages:
            t = page.extract_text() or ""
            # Bersihkan format teks PDF agar bisa dikenali pola regex
            t = re.sub(r"(\d)([A-Za-z])", r"\1 \2", t)
            t = re.sub(r"([a-z])([A-Z])", r"\1 \2", t)
            t = re.sub(r"([A-Za-z])(\d)", r"\1 \2", t)
            t = re.sub(r"\s{2,}", " ", t)
            text += t + "\n"

        if not text.strip():
            print("⚠️ Tidak ada teks terbaca dari PDF.")
            return jsonify({"status": "error", "message": "Tidak ada teks yang bisa dibaca dari file PDF ini."})

    except Exception as e:
        print("❌ ERROR saat membaca PDF:")
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"Gagal membaca file PDF: {str(e)}"})

    # --- 4️⃣ Ambil bagian daftar item antara 'Item' dan 'Info Pembayaran' ---
    match = re.search(r"# ?Item(.+?)Info ?Pembayaran", text, re.DOTALL | re.IGNORECASE)
    if not match:
        print("⚠️ Tidak ditemukan bagian daftar item di PDF.")
        return jsonify({"status": "error", "message": "Bagian daftar item tidak ditemukan di PDF."})

    daftar = match.group(1)

    # --- 5️⃣ Regex parsing baris item (dengan satuan opsional) ---
    pattern = re.compile(
        r"^\s*(\d+)\s+([A-Za-z0-9\s]+?)\s+(?:([A-Za-z]+)\s+)?([\d.,]+)\s+Rp\s*([\d.,]+)\s+Rp\s*([\d.,]+)",
        re.IGNORECASE | re.MULTILINE
    )

    def parse_num(val):
        val = val.replace("Rp", "").replace(" ", "").replace(",", "")
        if val.count(".") > 1:
            val = val.replace(".", "", val.count(".") - 1)
        try:
            return float(val)
        except:
            return 0.0

    items = []
    for m in pattern.finditer(daftar):
        satuan = m.group(3).strip() if m.group(3) else "-"
        items.append({
            "no": int(m.group(1)),
            "nama_barang": m.group(2).strip(),
            "satuan": satuan,
            "jumlah": parse_num(m.group(4)),
            "harga_satuan": parse_num(m.group(5)),
            "total_harga": parse_num(m.group(6))
        })

    # --- 6️⃣ Validasi hasil parsing ---
    if not items:
        print("⚠️ Tidak ada item yang berhasil diparsing dari PDF.")
        return jsonify({"status": "error", "message": "Tidak ada item yang berhasil dibaca dari PDF."})

    # Urutkan A-Z untuk konsistensi tampilan
    items.sort(key=lambda x: x["nama_barang"].lower())

    print(f"✅ Berhasil membaca {len(items)} item dari PDF.")
    return jsonify({"status": "ok", "data": items})

@app.route("/export_penjualan_pdf/<nomor_faktur>")
def export_penjualan_pdf(nomor_faktur):
    import io
    from datetime import datetime
 
    # === Buat buffer dan style dasar dulu ===
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Header", fontSize=11, alignment=1, leading=13))
    styles.add(ParagraphStyle(name="NormalLeft", fontSize=10, alignment=0, leading=12))
    styles.add(ParagraphStyle(name="BoldCenter", fontSize=10, alignment=1, leading=12))

    # === Ambil setting layout (dari tabel LayoutSetting) ===
    conn2 = sqlite3.connect(DB_PATH)
    conn2.row_factory = sqlite3.Row
    cur2 = conn2.cursor()
    cur2.execute("SELECT * FROM LayoutSetting LIMIT 1")
    setting = cur2.fetchone()
    conn2.close()

    if setting:
        header_align = {"LEFT": 0, "CENTER": 1, "RIGHT": 2}.get(setting["posisi_header"], 1)
        font_size = setting["font_size_header"]
        leading = setting["jarak_header"]
        styles.add(ParagraphStyle(name="HeaderDynamic", fontSize=font_size, alignment=header_align, leading=leading))
    else:
        # fallback jika setting belum ada
        styles.add(ParagraphStyle(name="HeaderDynamic", fontSize=11, alignment=1, leading=13))

    # === Ambil data header dari Penjualan ===
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT id, tanggal_jual, konsumen, nomor_faktur_beli
        FROM Penjualan
        WHERE nomor_faktur_beli = ?
    """, (nomor_faktur,))
    penjualan = cur.fetchone()

    if not penjualan:
        conn.close()
        return f"❌ Faktur {nomor_faktur} tidak ditemukan di tabel Penjualan.", 404

    penjualan_id = penjualan["id"]

    # === Ambil item barang ===
    cur.execute("""
        SELECT nama_barang, satuan, jumlah, harga_jual
        FROM ItemPenjualan
        WHERE penjualan_id = ?
    """, (penjualan_id,))
    items = [dict(r) for r in cur.fetchall()]
    conn.close()

    if not items:
        return f"⚠️ Tidak ada item untuk faktur {nomor_faktur}.", 404

    # === Mulai isi konten PDF ===
    content = []

    # === Header koperasi (diambil dari LayoutSetting) ===
    if setting:
        koperasi_info = [
            setting["nama_koperasi"],
            setting["alamat"],
            setting["kontak"]
        ]
    else:
        koperasi_info = [
            "KOPERASI KONSUMEN MARSUKI RAGIL MANDIRI",
            "Jl Ryacudu gg Surya Alam no 09 Korpri Raya, Kec. Sukarame Kota Bandar Lampung",
            "081262014034"
        ]

    for line in koperasi_info:
        content.append(Paragraph(line, styles["HeaderDynamic"]))

    content.append(Spacer(1, 10))

    # === Info faktur ===
    content.append(Paragraph("<b>FAKTUR</b>", styles["BoldCenter"]))
    content.append(Paragraph(f"<b>#{nomor_faktur}</b>", styles["BoldCenter"]))
    content.append(Spacer(1, 8))
    content.append(Paragraph(f"Tanggal: {penjualan['tanggal_jual']}", styles["NormalLeft"]))
    content.append(Paragraph(f"Tagih Kepada: {penjualan['konsumen']}", styles["NormalLeft"]))
    content.append(Spacer(1, 10))

    # === Tabel item ===
    table_data = [["#", "Item", "Unit", "Kuantitas", "Biaya Satuan", "Total"]]
    subtotal = 0
    for i, item in enumerate(items, 1):
        total = item["jumlah"] * item["harga_jual"]
        subtotal += total
        table_data.append([
            str(i),
            item["nama_barang"],
            item["satuan"],
            f"{item['jumlah']:,.0f}",
            f"Rp{item['harga_jual']:,.0f}",
            f"Rp{total:,.0f}"
        ])

    table = Table(table_data, colWidths=[1*cm, 6*cm, 2*cm, 2.5*cm, 3*cm, 3*cm])
    table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("ALIGN", (1,1), (1,-1), "LEFT"),
        ("ALIGN", (4,1), (-1,-1), "RIGHT"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
    ]))
    content.append(table)
    content.append(Spacer(1, 10))

    # === Info pembayaran ===
    content.append(Paragraph("<b>Info Pembayaran</b>", styles["NormalLeft"]))
    content.append(Paragraph("Bri 579601000052568", styles["NormalLeft"]))
    content.append(Paragraph("An koperasi konsumen marsuki ragil mandiri", styles["NormalLeft"]))
    content.append(Paragraph("Bsi 7307554806", styles["NormalLeft"]))
    content.append(Paragraph("An/ koperasi marsuki ragil mandiri", styles["NormalLeft"]))
    content.append(Spacer(1, 8))

    # === Subtotal & Total ===
    total_str = f"Rp{subtotal:,.0f}"
    content.append(Paragraph(f"<b>Subtotal:</b> {total_str}", styles["NormalLeft"]))
    content.append(Paragraph(f"<b>Total:</b> {total_str}", styles["NormalLeft"]))
    content.append(Paragraph(f"<b>Saldo Terutang:</b> {total_str}", styles["NormalLeft"]))
    content.append(Spacer(1, 12))

    # === Footer ===
    content.append(Paragraph(
        "KOPERASI KONSUMEN MARSUKI RAGIL MANDIRI<br/>"
        "KELURAHAN KORPRI RAYA KECAMATAN SUKARAME KOTA BANDAR LAMPUNG",
        styles["BoldCenter"]
    ))
    content.append(Paragraph(datetime.now().strftime("%d/%m/%Y"), styles["BoldCenter"]))

    # === Build dokumen ===
    doc.build(content)
    buffer.seek(0)

    filename = f"{nomor_faktur}_{datetime.now().strftime('%d_%m_%Y')}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")

@app.route("/debug_pdf", methods=["POST"])
def debug_pdf():
    from PyPDF2 import PdfReader

    if "file" not in request.files:
        return jsonify({"status": "error", "message": "File tidak ditemukan"})

    file = request.files["file"]
    reader = PdfReader(file)
    text = ""
    for page in reader.pages:
        text += page.extract_text() + "\n"

    # Ambil hanya 50 baris pertama biar nggak kepanjangan
    lines = [ln for ln in text.splitlines() if ln.strip()]
    preview = "\n".join(lines[:50])

    return jsonify({"status": "ok", "preview": preview})

@app.route("/export_penjualan_excel_template/<nomor_faktur>")
def export_penjualan_excel_template(nomor_faktur):
    import os, io
    from datetime import datetime
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.cell.cell import MergedCell

    # === Helper aman untuk tulis ke merged cell ===
    def write_safe(ws, cell_ref, value, align=None):
        cell = ws[cell_ref]
        if isinstance(cell, MergedCell):
            for merged in ws.merged_cells.ranges:
                if cell_ref in merged:
                    top_left = merged.coord.split(":")[0]
                    ws[top_left].value = value
                    if align:
                        ws[top_left].alignment = align
                    return
        else:
            ws[cell_ref].value = value
            if align:
                ws[cell_ref].alignment = align

    # === Pastikan template Excel tersedia ===
    base_path = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_path, "Master Kops.xlsx")

    if not os.path.exists(template_path):
        return f"❌ Template Excel tidak ditemukan: {template_path}", 500

    # === Ambil data dari database ===
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("SELECT * FROM Penjualan WHERE nomor_faktur_beli = ?", (nomor_faktur,))
    penjualan = cur.fetchone()

    if not penjualan:
        conn.close()
        return f"❌ Faktur {nomor_faktur} tidak ditemukan.", 404

    cur.execute("""
        SELECT nama_barang, satuan, jumlah, harga_jual
        FROM ItemPenjualan
        WHERE penjualan_id = ?
    """, (penjualan["id"],))
    rows = cur.fetchall()
    conn.close()

    # === GROUPING (Gabungkan nama barang yang sama) ===
    grouped = {}

    for r in rows:
        key = (r["nama_barang"].strip().lower(), r["satuan"].strip().lower())

        if key not in grouped:
            grouped[key] = {
                "nama_barang": r["nama_barang"],
                "satuan": r["satuan"],
                "jumlah": float(r["jumlah"]),
                "harga_jual": float(r["harga_jual"])
            }
        else:
            # qty ditambah
            grouped[key]["jumlah"] += float(r["jumlah"])

            # jika harga jual beda, tetap ambil harga jual terakhir yang dimasukkan
            grouped[key]["harga_jual"] = float(r["harga_jual"])

    items = list(grouped.values())

    # === Load workbook ===
    wb = load_workbook(template_path)
    ws = wb.active

    # === Header ===
    write_safe(ws, "F6", penjualan["nomor_faktur_beli"])
    write_safe(ws, "F7", penjualan["konsumen"])
    write_safe(ws, "F8", penjualan["tanggal_jual"])

    TEMPLATE_ROW = 9
    jumlah_items = len(items)

    # Geser gambar
    shift = max(0, jumlah_items - 1)
    if shift > 0:
        for img in ws._images:
            img.anchor._from.row += shift

    # Insert rows
    if jumlah_items > 1:
        ws.insert_rows(TEMPLATE_ROW + 1, jumlah_items - 1)

    # === Isi item ===
    for i, item in enumerate(items):
        row = TEMPLATE_ROW + i

        write_safe(ws, f"B{row}", i + 1)
        write_safe(ws, f"C{row}", item["nama_barang"])
        write_safe(ws, f"D{row}", item["satuan"], Alignment(horizontal="center"))
        write_safe(ws, f"E{row}", item["jumlah"], Alignment(horizontal="center"))
        write_safe(ws, f"F{row}", item["harga_jual"], Alignment(horizontal="right"))
        write_safe(ws, f"G{row}", item["jumlah"] * item["harga_jual"], Alignment(horizontal="right"))

    # === Total ===
    TOTAL_ROW = TEMPLATE_ROW + jumlah_items + 1
    total_semua = sum(i["jumlah"] * i["harga_jual"] for i in items)

    write_safe(ws, f"F{TOTAL_ROW}", "TOTAL", Alignment(horizontal="right"))
    write_safe(ws, f"G{TOTAL_ROW}", total_semua, Alignment(horizontal="right"))

    # === Output ===
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{nomor_faktur}_Blanko_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# === SIMULASI HARGA JUAL (TIDAK SIMPAN DATABASE) ===
@app.route("/simulasi_harga_jual")
def simulasi_harga_jual():
    return render_template(
        "simulasi_harga_jual.html",
        today=date.today().strftime("%Y-%m-%d")
    )

# === API: HITUNG SIMULASI HARGA JUAL ===
@app.route("/api/simulasi_harga_jual")
def api_simulasi_harga_jual():
    nama = request.args.get("nama", "").strip()
    satuan = request.args.get("satuan", "").strip()
    qty = float(request.args.get("qty", 1) or 1)

    if not nama:
        return jsonify({"harga_beli": 0, "harga_jual": 0, "total": 0})

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # 1️⃣ ambil harga jual terakhir
    cur.execute("""
        SELECT harga_jual, harga_beli
        FROM ItemPenjualan
        WHERE LOWER(nama_barang)=LOWER(?) AND satuan=?
        ORDER BY id DESC LIMIT 1
    """, (nama, satuan))
    row = cur.fetchone()

    if row:
        harga_jual = row[0]
        harga_beli = row[1]
    else:
        # 2️⃣ fallback: harga beli terakhir dari pembelian
        cur.execute("""
            SELECT harga_satuan
            FROM ItemBelanja
            WHERE LOWER(nama_barang)=LOWER(?) AND satuan=?
            ORDER BY id DESC LIMIT 1
        """, (nama, satuan))
        hb = cur.fetchone()
        harga_beli = hb[0] if hb else 0
        harga_jual = round(harga_beli * 1.2) if harga_beli else 0

    conn.close()

    total = qty * harga_jual

    return jsonify({
        "harga_beli": harga_beli,
        "harga_jual": harga_jual,
        "total": total
    })

# ============================================================
# === PENJUALAN RAB — Skema baru, terpisah dari penjualan lama ===
# ============================================================

@app.route("/riwayat_barang")
def riwayat_barang_page():
    return render_template("riwayat_barang.html")

@app.route("/progress_barang")
def progress_barang_page():
    return render_template("progress_barang.html")

@app.route("/api/progress_barang")
def api_progress_barang():
    nama = request.args.get("nama", "").strip()
    if not nama:
        return jsonify({"status": "error", "message": "Nama barang kosong"})

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Per tanggal: qty, harga rata-rata tertimbang, total nilai
    cur.execute("""
        SELECT n.tanggal AS tgl,
               n.supplier AS supplier,
               n.nomor_faktur AS faktur,
               SUM(i.jumlah) AS qty,
               ROUND(
                   SUM(COALESCE(i.total_harga, i.jumlah * i.harga_satuan))
                   / NULLIF(SUM(i.jumlah), 0)
               , 0) AS harga_avg,
               SUM(COALESCE(i.total_harga, i.jumlah * i.harga_satuan)) AS nilai,
               MAX(i.satuan) AS satuan
        FROM ItemBelanja i JOIN NotaBelanja n ON i.nota_id = n.id
        WHERE LOWER(i.nama_barang) = LOWER(?)
        GROUP BY n.tanggal, n.id
        ORDER BY n.tanggal
    """, (nama,))
    rows = cur.fetchall()
    conn.close()

    if not rows:
        return jsonify({
            "status": "ok", "nama": nama, "satuan": "",
            "dates": [], "qty": [], "harga": [], "nilai": [],
            "labels": [],
            "totals": {"qty": 0, "nilai": 0, "count": 0,
                       "harga_min": 0, "harga_max": 0, "harga_avg": 0}
        })

    satuan      = next((r["satuan"] for r in rows if r["satuan"]), "")
    dates       = [r["tgl"]       for r in rows]
    qty_list    = [r["qty"]   or 0 for r in rows]
    harga_list  = [r["harga_avg"] or 0 for r in rows]
    nilai_list  = [r["nilai"] or 0 for r in rows]
    labels      = [f"{r['tgl']} | {r['supplier'] or ''} | {r['faktur'] or ''}" for r in rows]

    total_qty   = sum(qty_list)
    total_nilai = sum(nilai_list)
    valid_h     = [h for h in harga_list if h > 0]

    return jsonify({
        "status": "ok", "nama": nama, "satuan": satuan,
        "dates":  dates,
        "qty":    qty_list,
        "harga":  harga_list,
        "nilai":  nilai_list,
        "labels": labels,
        "totals": {
            "qty":       total_qty,
            "nilai":     total_nilai,
            "count":     len(rows),
            "harga_min": min(valid_h) if valid_h else 0,
            "harga_max": max(valid_h) if valid_h else 0,
            "harga_avg": round(sum(valid_h) / len(valid_h)) if valid_h else 0,
        }
    })

@app.route("/api/daftar_barang")
def api_daftar_barang():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT i.nama_barang, i.satuan, i.harga_satuan AS harga_beli, n.tanggal AS tanggal_beli
        FROM ItemBelanja i
        JOIN NotaBelanja n ON i.nota_id = n.id
        WHERE i.nama_barang IS NOT NULL AND i.nama_barang <> ''
        ORDER BY n.tanggal DESC, i.id DESC
    """)
    rows = cur.fetchall()

    cur.execute("SELECT DISTINCT LOWER(nama_barang) AS nm FROM ItemPenjualan WHERE nama_barang IS NOT NULL AND nama_barang <> ''")
    jual_names = {r['nm'] for r in cur.fetchall()}

    cur.execute("SELECT DISTINCT LOWER(nama_barang_jual) AS nm FROM ItemPenjualanRAB WHERE nama_barang_jual IS NOT NULL AND nama_barang_jual <> ''")
    rab_names = {r['nm'] for r in cur.fetchall()}

    conn.close()

    seen = {}
    for r in rows:
        key = r['nama_barang'].lower().strip()
        if key not in seen:
            seen[key] = {
                "nama": r["nama_barang"],
                "satuan": r["satuan"],
                "harga_beli": r["harga_beli"],
                "tanggal_beli": r["tanggal_beli"],
                "ada_jual": key in jual_names,
                "ada_rab": key in rab_names,
            }

    return jsonify({"status": "ok", "items": list(seen.values())})

@app.route("/transaksi_rab")
def transaksi_rab():
    return render_template("transaksi_rab.html", today=date.today().strftime("%Y-%m-%d"))


@app.route("/api/transaksi_rab/simpan", methods=["POST"])
def simpan_transaksi_rab():
    """Simpan transaksi gabungan: beli ke NotaBelanja/ItemBelanja + jual ke PenjualanRAB/ItemPenjualanRAB."""
    data = request.get_json()
    if not data:
        return jsonify({"status": "error", "message": "Data kosong"})

    tanggal_beli  = data.get("tanggal_beli",  "").strip()
    supplier      = data.get("supplier",       "").strip()
    nomor_faktur  = data.get("nomor_faktur",   "").strip()
    tanggal_jual  = data.get("tanggal_jual",   "").strip()
    konsumen      = data.get("konsumen",       "").strip()
    catatan       = data.get("catatan",        "").strip()
    items         = data.get("items",          [])

    if not tanggal_beli or not nomor_faktur or not supplier:
        return jsonify({"status": "error", "message": "Tanggal beli, nomor faktur, dan supplier wajib diisi."})
    if not tanggal_jual or not konsumen:
        return jsonify({"status": "error", "message": "Tanggal jual dan konsumen wajib diisi."})
    if not items:
        return jsonify({"status": "error", "message": "Minimal 1 item harus diisi."})

    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()
    try:
        # ── 1. Simpan NotaBelanja ──
        cur.execute("SELECT id FROM NotaBelanja WHERE nomor_faktur = ?", (nomor_faktur,))
        existing_nota = cur.fetchone()
        if existing_nota:
            cur.execute("DELETE FROM ItemBelanja WHERE nota_id = ?", (existing_nota[0],))
            cur.execute("DELETE FROM NotaBelanja WHERE id = ?", (existing_nota[0],))

        cur.execute("""
            INSERT INTO NotaBelanja (tanggal, supplier, nomor_faktur)
            VALUES (?, ?, ?)
        """, (tanggal_beli, supplier, nomor_faktur))
        nota_id = cur.lastrowid

        # ── 2. Simpan ItemBelanja ──
        saved_beli = 0
        for it in items:
            nama_beli   = str(it.get("nama_beli",   "")).strip()
            jumlah_beli = float(it.get("jumlah_beli", 0) or 0)
            satuan_beli = str(it.get("satuan_beli",  "")).strip()
            harga_beli  = float(it.get("harga_beli",  0) or 0)
            if not nama_beli:
                continue
            total_harga = jumlah_beli * harga_beli
            cur.execute("""
                INSERT INTO ItemBelanja (nota_id, nama_barang, jumlah, satuan, harga_satuan, total_harga)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (nota_id, nama_beli, jumlah_beli, satuan_beli, harga_beli, total_harga))
            saved_beli += 1

        # ── 3. Simpan PenjualanRAB ──
        cur.execute("SELECT id FROM PenjualanRAB WHERE nomor_faktur_beli = ?", (nomor_faktur,))
        existing_rab = cur.fetchone()
        if existing_rab:
            cur.execute("DELETE FROM ItemPenjualanRAB WHERE penjualan_rab_id = ?", (existing_rab[0],))
            cur.execute("DELETE FROM PenjualanRAB WHERE id = ?", (existing_rab[0],))

        cur.execute("""
            INSERT INTO PenjualanRAB (tanggal_jual, konsumen, nomor_faktur_beli, catatan)
            VALUES (?, ?, ?, ?)
        """, (tanggal_jual, konsumen, nomor_faktur, catatan))
        pjid = cur.lastrowid

        # ── 4. Simpan ItemPenjualanRAB ──
        saved_jual = 0
        for it in items:
            nama_jual   = str(it.get("nama_jual",   "")).strip()
            jumlah_jual = float(it.get("jumlah_jual", 0) or 0)
            satuan_jual = str(it.get("satuan_jual",  "")).strip()
            harga_jual  = float(it.get("harga_jual",  0) or 0)
            nama_beli   = str(it.get("nama_beli",   "")).strip()
            jumlah_beli = float(it.get("jumlah_beli", 0) or 0)
            satuan_beli = str(it.get("satuan_beli",  "")).strip()
            harga_beli  = float(it.get("harga_beli",  0) or 0)
            if not nama_jual:
                continue
            total_jual = jumlah_jual * harga_jual
            cur.execute("""
                INSERT INTO ItemPenjualanRAB
                (penjualan_rab_id, nama_barang_beli, jumlah_beli, satuan_beli, harga_beli,
                 nama_barang_jual, jumlah_jual, satuan_jual, harga_jual, total_jual)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (pjid, nama_beli, jumlah_beli, satuan_beli, harga_beli,
                  nama_jual, jumlah_jual, satuan_jual, harga_jual, total_jual))
            saved_jual += 1

        if saved_beli == 0 and saved_jual == 0:
            conn.rollback()
            return jsonify({"status": "error", "message": "Tidak ada item valid yang dapat disimpan."})

        conn.commit()
        return jsonify({
            "status": "ok",
            "message": f"Tersimpan: {saved_beli} item beli, {saved_jual} item jual RAB.",
            "nota_id": nota_id,
            "rab_id": pjid
        })
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": str(e)})
    finally:
        conn.close()


@app.route("/api/transaksi_rab/list")
def list_transaksi_rab():
    """Daftar semua transaksi RAB (beli+jual) yang sudah disimpan."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT n.nomor_faktur, n.tanggal AS tanggal_beli, n.supplier,
               p.tanggal_jual, p.konsumen,
               COUNT(i.id)    AS jumlah_item,
               SUM(i.total_jual) AS grand_total
        FROM NotaBelanja n
        JOIN PenjualanRAB p ON p.nomor_faktur_beli = n.nomor_faktur
        LEFT JOIN ItemPenjualanRAB i ON i.penjualan_rab_id = p.id
        GROUP BY n.id
        ORDER BY n.id DESC
        LIMIT 100
    """)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"status": "ok", "data": rows})


@app.route("/api/transaksi_rab/get/<nomor_faktur>")
def get_transaksi_rab(nomor_faktur):
    """Ambil data lengkap transaksi RAB (beli+jual) berdasarkan nomor faktur."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("SELECT * FROM NotaBelanja WHERE nomor_faktur = ?", (nomor_faktur,))
    nota = cur.fetchone()
    if not nota:
        conn.close()
        return jsonify({"status": "not_found"})

    cur.execute("SELECT * FROM PenjualanRAB WHERE nomor_faktur_beli = ?", (nomor_faktur,))
    pjual = cur.fetchone()
    if not pjual:
        conn.close()
        return jsonify({"status": "not_found"})

    cur.execute("SELECT * FROM ItemPenjualanRAB WHERE penjualan_rab_id = ?", (pjual["id"],))
    items = [dict(r) for r in cur.fetchall()]
    conn.close()

    return jsonify({
        "status":       "ok",
        "tanggal_beli": nota["tanggal"],
        "supplier":     nota["supplier"],
        "nomor_faktur": nomor_faktur,
        "tanggal_jual": pjual["tanggal_jual"],
        "konsumen":     pjual["konsumen"],
        "catatan":      pjual["catatan"] or "",
        "items":        items,
    })


@app.route("/penjualan_rab")
def penjualan_rab():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT nomor_faktur, supplier FROM NotaBelanja ORDER BY id DESC LIMIT 100")
    nota_list = [dict(r) for r in cur.fetchall()]
    conn.close()
    return render_template("penjualan_rab.html",
                           nota_list=nota_list,
                           today=date.today().strftime("%Y-%m-%d"))


@app.route("/api/penjualan_rab/simpan", methods=["POST"])
def simpan_penjualan_rab():
    """Simpan transaksi penjualan RAB ke tabel baru."""
    data = request.get_json()
    if not data:
        return jsonify({"status": "error", "message": "Data kosong"})

    tanggal        = data.get("tanggal", "").strip()
    konsumen       = data.get("konsumen", "").strip()
    nomor_faktur   = data.get("nomor_faktur", "").strip()
    catatan        = data.get("catatan", "").strip()
    items          = data.get("items", [])

    if not tanggal or not konsumen or not nomor_faktur:
        return jsonify({"status": "error", "message": "Tanggal, konsumen, dan nomor faktur wajib diisi."})
    if not items:
        return jsonify({"status": "error", "message": "Minimal 1 item harus diisi."})

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    try:
        # Cek apakah faktur ini sudah ada — jika ya, update (hapus lama, insert baru)
        cur.execute("SELECT id FROM PenjualanRAB WHERE nomor_faktur_beli = ?", (nomor_faktur,))
        existing = cur.fetchone()
        if existing:
            cur.execute("DELETE FROM ItemPenjualanRAB WHERE penjualan_rab_id = ?", (existing[0],))
            cur.execute("DELETE FROM PenjualanRAB WHERE id = ?", (existing[0],))

        cur.execute("""
            INSERT INTO PenjualanRAB (tanggal_jual, konsumen, nomor_faktur_beli, catatan)
            VALUES (?, ?, ?, ?)
        """, (tanggal, konsumen, nomor_faktur, catatan))
        pjid = cur.lastrowid

        saved = 0
        for it in items:
            nama_jual  = str(it.get("nama_jual", "")).strip()
            jumlah_jual= float(it.get("jumlah_jual", 0) or 0)
            satuan_jual= str(it.get("satuan_jual", "")).strip()
            harga_jual = float(it.get("harga_jual", 0) or 0)
            # Simpan semua item yang punya nama_jual, walaupun harga_jual belum diisi (0)
            if not nama_jual:
                continue
            total = jumlah_jual * harga_jual
            cur.execute("""
                INSERT INTO ItemPenjualanRAB
                (penjualan_rab_id, nama_barang_beli, jumlah_beli, satuan_beli, harga_beli,
                 nama_barang_jual, jumlah_jual, satuan_jual, harga_jual, total_jual)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                pjid,
                it.get("nama_beli", ""), it.get("jumlah_beli", 0),
                it.get("satuan_beli", ""), it.get("harga_beli", 0),
                nama_jual, jumlah_jual, satuan_jual, harga_jual, total
            ))
            saved += 1

        if saved == 0:
            conn.rollback()
            return jsonify({"status": "error", "message": "Tidak ada item valid yang tersimpan (nama barang kosong semua)."})

        conn.commit()
        return jsonify({"status": "ok", "message": f"{saved} item berhasil disimpan.", "id": pjid})
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": str(e)})
    finally:
        conn.close()


@app.route("/api/penjualan_rab/get/<nomor_faktur>")
def get_penjualan_rab(nomor_faktur):
    """Ambil data penjualan RAB yang sudah tersimpan (untuk load/edit ulang)."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM PenjualanRAB WHERE nomor_faktur_beli = ?", (nomor_faktur,))
    header = cur.fetchone()
    if not header:
        conn.close()
        return jsonify({"status": "not_found"})
    cur.execute("SELECT * FROM ItemPenjualanRAB WHERE penjualan_rab_id = ?", (header["id"],))
    items = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"status": "ok", "header": dict(header), "items": items})


@app.route("/api/penjualan_rab/list")
def list_penjualan_rab():
    """Daftar semua transaksi PenjualanRAB."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT p.id, p.tanggal_jual, p.konsumen, p.nomor_faktur_beli, p.dibuat_pada,
               COUNT(i.id) as jumlah_item,
               SUM(i.total_jual) as grand_total
        FROM PenjualanRAB p
        LEFT JOIN ItemPenjualanRAB i ON p.id = i.penjualan_rab_id
        GROUP BY p.id
        ORDER BY p.id DESC LIMIT 50
    """)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"status": "ok", "data": rows})


@app.route("/api/penjualan_rab/hapus/<int:pj_id>", methods=["POST"])
def hapus_penjualan_rab(pj_id):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("DELETE FROM ItemPenjualanRAB WHERE penjualan_rab_id = ?", (pj_id,))
    cur.execute("DELETE FROM PenjualanRAB WHERE id = ?", (pj_id,))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok"})


# ============================================================
# === EXPORT PDF — PenjualanRAB ===
# ============================================================
@app.route("/export_rab_pdf/<nomor_faktur>")
def export_rab_pdf(nomor_faktur):
    import io
    from datetime import datetime

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="NormalLeft",  fontSize=10, alignment=0, leading=12))
    styles.add(ParagraphStyle(name="BoldCenter",  fontSize=10, alignment=1, leading=12, fontName="Helvetica-Bold"))

    # Ambil LayoutSetting
    conn2 = sqlite3.connect(DB_PATH)
    conn2.row_factory = sqlite3.Row
    cur2 = conn2.cursor()
    cur2.execute("SELECT * FROM LayoutSetting LIMIT 1")
    setting = cur2.fetchone()
    conn2.close()

    if setting:
        header_align = {"LEFT": 0, "CENTER": 1, "RIGHT": 2}.get(setting["posisi_header"], 1)
        font_size    = int(setting["font_size_header"] or 11)
        leading      = float(setting["jarak_header"] or 13)
    else:
        header_align, font_size, leading = 1, 11, 13

    styles.add(ParagraphStyle(name="HeaderDynamic", fontSize=font_size,
                              alignment=header_align, leading=leading))

    # Ambil data PenjualanRAB
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur  = conn.cursor()

    cur.execute("SELECT * FROM PenjualanRAB WHERE nomor_faktur_beli = ?", (nomor_faktur,))
    pj = cur.fetchone()
    if not pj:
        conn.close()
        return f"❌ Faktur RAB {nomor_faktur} tidak ditemukan.", 404

    cur.execute("""
        SELECT nama_barang_jual AS nama_barang,
               satuan_jual      AS satuan,
               jumlah_jual      AS jumlah,
               harga_jual,
               total_jual
        FROM ItemPenjualanRAB
        WHERE penjualan_rab_id = ?
        ORDER BY id ASC
    """, (pj["id"],))
    raw_items = [dict(r) for r in cur.fetchall()]
    conn.close()

    if not raw_items:
        return f"❌ Tidak ada item untuk faktur RAB {nomor_faktur}.", 404

    # Grouping: nama+satuan sama dijumlah qty (TIMUN 60+3=63)
    _g = {}
    for r in raw_items:
        if not r.get("nama_barang"): continue
        k = (str(r["nama_barang"]).strip().lower(), str(r.get("satuan") or "").strip().lower())
        if k not in _g:
            _g[k] = dict(r)
            _g[k]["jumlah"] = float(r["jumlah"] or 0)
        else:
            _g[k]["jumlah"]    += float(r["jumlah"] or 0)
            _g[k]["harga_jual"] = float(r["harga_jual"] or 0)
            _g[k]["total_jual"] = _g[k]["jumlah"] * _g[k]["harga_jual"]
    items = list(_g.values())

    if not items:
        return f"⚠️ Tidak ada item untuk faktur RAB {nomor_faktur}.", 404

    # Header koperasi
    kop = (setting["nama_koperasi"], setting["alamat"], setting["kontak"]) if setting else (
        "KOPERASI KONSUMEN MARSUKI RAGIL MANDIRI",
        "Jl Ryacudu gg Surya Alam no 09 Korpri Raya, Kec. Sukarame Kota Bandar Lampung",
        "081262014034"
    )

    content = []
    for line in kop:
        if line:
            content.append(Paragraph(line, styles["HeaderDynamic"]))
    content.append(Spacer(1, 10))

    content.append(Paragraph("<b>FAKTUR</b>", styles["BoldCenter"]))
    content.append(Paragraph(f"<b>#{nomor_faktur}</b>", styles["BoldCenter"]))
    content.append(Spacer(1, 8))
    content.append(Paragraph(f"Tanggal  : {pj['tanggal_jual']}", styles["NormalLeft"]))
    content.append(Paragraph(f"Kepada   : {pj['konsumen']}", styles["NormalLeft"]))
    if pj["catatan"]:
        content.append(Paragraph(f"Catatan  : {pj['catatan']}", styles["NormalLeft"]))
    content.append(Spacer(1, 10))

    # Tabel item — pakai data JUAL
    table_data = [["#", "Nama Barang", "Sat", "Qty", "Harga Satuan", "Total"]]
    subtotal = 0
    for i, it in enumerate(items, 1):
        total = float(it["jumlah"] or 0) * float(it["harga_jual"] or 0)
        subtotal += total
        table_data.append([
            str(i),
            it["nama_barang"] or "-",
            it["satuan"] or "-",
            f"{float(it['jumlah']):,.0f}",
            f"Rp{float(it['harga_jual']):,.0f}",
            f"Rp{total:,.0f}"
        ])

    tbl_color = colors.lightgrey
    if setting and setting["table_header_color"]:
        try:
            tbl_color = getattr(colors, setting["table_header_color"], colors.lightgrey)
        except Exception:
            pass

    tbl = Table(table_data, colWidths=[0.8*cm, 6.2*cm, 1.8*cm, 2*cm, 3*cm, 3*cm])
    tbl.setStyle(TableStyle([
        ("GRID",       (0,0), (-1,-1), 0.5, colors.black),
        ("BACKGROUND", (0,0), (-1, 0), tbl_color),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("ALIGN",      (1,1), (1, -1), "LEFT"),
        ("ALIGN",      (4,1), (-1,-1), "RIGHT"),
        ("FONTNAME",   (0,0), (-1, 0), "Helvetica-Bold"),
    ]))
    content.append(tbl)
    content.append(Spacer(1, 10))

    # Info pembayaran
    content.append(Paragraph("<b>Info Pembayaran</b>", styles["NormalLeft"]))
    content.append(Paragraph("BRI 579601000052568 — An. Koperasi Konsumen Marsuki Ragil Mandiri", styles["NormalLeft"]))
    content.append(Paragraph("BSI 7307554806 — An. Koperasi Marsuki Ragil Mandiri", styles["NormalLeft"]))
    content.append(Spacer(1, 8))

    total_str = f"Rp{subtotal:,.0f}"
    content.append(Paragraph(f"<b>Subtotal      :</b> {total_str}", styles["NormalLeft"]))
    content.append(Paragraph(f"<b>Total         :</b> {total_str}", styles["NormalLeft"]))
    content.append(Paragraph(f"<b>Saldo Terutang:</b> {total_str}", styles["NormalLeft"]))
    content.append(Spacer(1, 14))

    footer_align_val = {"LEFT": 0, "CENTER": 1, "RIGHT": 2}.get(
        setting["footer_align"] if setting else "CENTER", 1)
    styles.add(ParagraphStyle(name="Footer", fontSize=10,
                              alignment=footer_align_val, leading=12))
    content.append(Paragraph(
        f"{kop[0]}<br/>{kop[1]}",
        styles["Footer"]))
    content.append(Paragraph(datetime.now().strftime("%d/%m/%Y"), styles["Footer"]))

    doc.build(content)
    buffer.seek(0)

    filename = f"RAB_{nomor_faktur}_{datetime.now().strftime('%d_%m_%Y')}.pdf"
    return send_file(buffer, as_attachment=True,
                     download_name=filename, mimetype="application/pdf")


# ============================================================
# === EXPORT EXCEL LAPORAN — PenjualanRAB ===
# ============================================================
@app.route("/export_rab_excel/<nomor_faktur>")
def export_rab_excel(nomor_faktur):
    import io, pandas as pd
    from datetime import datetime

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur  = conn.cursor()

    cur.execute("SELECT * FROM PenjualanRAB WHERE nomor_faktur_beli = ?", (nomor_faktur,))
    pj = cur.fetchone()
    if not pj:
        conn.close()
        return "❌ Faktur RAB tidak ditemukan.", 404

    cur.execute("""
        SELECT
            nama_barang_beli  AS 'Nama Beli',
            jumlah_beli       AS 'Qty Beli',
            satuan_beli       AS 'Satuan Beli',
            harga_beli        AS 'Harga Beli',
            (jumlah_beli * harga_beli) AS 'Total Beli',
            nama_barang_jual  AS 'Nama Jual',
            jumlah_jual       AS 'Qty Jual',
            satuan_jual       AS 'Satuan Jual',
            harga_jual        AS 'Harga Jual',
            total_jual        AS 'Total Jual'
        FROM ItemPenjualanRAB
        WHERE penjualan_rab_id = ?
    """, (pj["id"],))
    rows = cur.fetchall()
    conn.close()
    # Export Excel — per item, tidak digrouping (Timun 60kg & Timun 3kg = 2 baris terpisah)

    if not rows:
        return "⚠️ Tidak ada item.", 404

    df = pd.DataFrame(rows)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        wb  = writer.book
        ws  = wb.add_worksheet("PenjualanRAB")
        writer.sheets["PenjualanRAB"] = ws

        # Formats
        title_fmt  = wb.add_format({"bold": True, "font_size": 13, "align": "center", "valign": "vcenter"})
        info_fmt   = wb.add_format({"bold": True})
        hdr_beli   = wb.add_format({"bold": True, "align": "center", "valign": "vcenter",
                                     "bg_color": "#CECBF6", "border": 1})
        hdr_jual   = wb.add_format({"bold": True, "align": "center", "valign": "vcenter",
                                     "bg_color": "#B7DEE8", "border": 1})
        text_fmt   = wb.add_format({"border": 1})
        money_fmt  = wb.add_format({"num_format": "#,##0", "align": "right", "border": 1})
        total_lbl  = wb.add_format({"bold": True, "align": "right",
                                     "bg_color": "#E2EFDA", "border": 1})
        total_val  = wb.add_format({"bold": True, "num_format": "#,##0",
                                     "align": "right", "bg_color": "#E2EFDA", "border": 1})

        # Judul
        ws.merge_range("A1:J1", "LAPORAN PENJUALAN RAB PER FAKTUR", title_fmt)
        ws.set_row(0, 22)

        # Info faktur
        ws.write("A3", "Nomor Faktur :", info_fmt); ws.write("B3", nomor_faktur)
        ws.write("A4", "Tanggal Jual :", info_fmt); ws.write("B4", pj["tanggal_jual"])
        ws.write("A5", "Konsumen     :", info_fmt); ws.write("B5", pj["konsumen"])
        if pj["catatan"]:
            ws.write("A6", "Catatan      :", info_fmt); ws.write("B6", pj["catatan"])

        # Header group — baris 7 (index 6)
        ws.merge_range(6, 0, 6, 4, "◀  DATA PEMBELIAN  ▶", hdr_beli)
        ws.merge_range(6, 5, 6, 9, "◀  DATA PENJUALAN  ▶", hdr_jual)
        ws.set_row(6, 18)

        # Header kolom — baris 8 (index 7)
        for ci, col in enumerate(df.columns):
            fmt = hdr_beli if ci < 5 else hdr_jual
            ws.write(7, ci, col, fmt)

        # Lebar kolom
        col_widths = [22, 8, 10, 13, 13, 22, 8, 10, 13, 13]
        for ci, w in enumerate(col_widths):
            ws.set_column(ci, ci, w)

        # Data
        money_cols = {"Harga Beli", "Total Beli", "Harga Jual", "Total Jual"}
        for ri, row in enumerate(df.values):
            for ci, val in enumerate(row):
                fmt = money_fmt if df.columns[ci] in money_cols else text_fmt
                ws.write(ri + 8, ci, val, fmt)

        # Total baris
        data_rows = len(df)
        tr = data_rows + 8   # excel row index (0-based)
        ws.write(tr, 3, "TOTAL BELI :", total_lbl)
        ws.write_formula(tr, 4,  f"=SUM(E9:E{tr})",   total_val)
        ws.write(tr, 8, "TOTAL JUAL :", total_lbl)
        ws.write_formula(tr, 9,  f"=SUM(J9:J{tr})",   total_val)
        ws.write(tr+1, 8, "LABA KOTOR :", total_lbl)
        ws.write_formula(tr+1, 9, f"=J{tr+1}-E{tr+1}", total_val)

    output.seek(0)
    filename = f"RAB_{nomor_faktur}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ============================================================
# === EXPORT BLANKO EXCEL (Template Master Kops) — PenjualanRAB ===
# ============================================================
@app.route("/export_rab_blanko/<nomor_faktur>")
def export_rab_blanko(nomor_faktur):
    import os, io
    from datetime import datetime
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.cell.cell import MergedCell

    def write_safe(ws, cell_ref, value, align=None):
        cell = ws[cell_ref]
        if isinstance(cell, MergedCell):
            for merged in ws.merged_cells.ranges:
                if cell_ref in merged:
                    top_left = merged.coord.split(":")[0]
                    ws[top_left].value = value
                    if align:
                        ws[top_left].alignment = align
                    return
        else:
            ws[cell_ref].value = value
            if align:
                ws[cell_ref].alignment = align

    base_path     = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_path, "Master Kops.xlsx")
    if not os.path.exists(template_path):
        return f"❌ Template Excel tidak ditemukan: {template_path}", 500

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur  = conn.cursor()

    cur.execute("SELECT * FROM PenjualanRAB WHERE nomor_faktur_beli = ?", (nomor_faktur,))
    pj = cur.fetchone()
    if not pj:
        conn.close()
        return f"❌ Faktur RAB {nomor_faktur} tidak ditemukan.", 404

    # Ambil data SISI JUAL — per item apa adanya, TIDAK digabung
    cur.execute("""
        SELECT nama_barang_jual AS nama_barang,
               satuan_jual      AS satuan,
               jumlah_jual      AS jumlah,
               harga_jual
        FROM ItemPenjualanRAB
        WHERE penjualan_rab_id = ?
        ORDER BY id ASC
    """, (pj["id"],))
    rows = cur.fetchall()
    conn.close()

    if not rows:
        return f"❌ Tidak ada item untuk faktur RAB {nomor_faktur}.", 404

    # Grouping: nama+satuan sama dijumlah qty (TIMUN 60+3=63)
    _gb = {}
    for r in rows:
        if not r["nama_barang"]: continue
        k = (str(r["nama_barang"]).strip().lower(), str(r["satuan"] or "").strip().lower())
        if k not in _gb:
            _gb[k] = {
                "nama_barang": r["nama_barang"],
                "satuan":      r["satuan"],
                "jumlah":      float(r["jumlah"] or 0),
                "harga_jual":  float(r["harga_jual"] or 0)
            }
        else:
            _gb[k]["jumlah"]    += float(r["jumlah"] or 0)
            _gb[k]["harga_jual"] = float(r["harga_jual"] or 0)
    items = list(_gb.values())

    wb = load_workbook(template_path)
    ws = wb.active

    write_safe(ws, "F6", pj["nomor_faktur_beli"])
    write_safe(ws, "F7", pj["konsumen"])
    write_safe(ws, "F8", pj["tanggal_jual"])

    TEMPLATE_ROW  = 9
    jumlah_items  = len(items)
    shift         = max(0, jumlah_items - 1)

    if shift > 0:
        for img in ws._images:
            img.anchor._from.row += shift

    if jumlah_items > 1:
        ws.insert_rows(TEMPLATE_ROW + 1, jumlah_items - 1)

    for i, item in enumerate(items):
        row = TEMPLATE_ROW + i
        write_safe(ws, f"B{row}", i + 1)
        write_safe(ws, f"C{row}", item["nama_barang"])
        write_safe(ws, f"D{row}", item["satuan"],   Alignment(horizontal="center"))
        write_safe(ws, f"E{row}", item["jumlah"],   Alignment(horizontal="center"))
        write_safe(ws, f"F{row}", item["harga_jual"], Alignment(horizontal="right"))
        write_safe(ws, f"G{row}", item["jumlah"] * item["harga_jual"], Alignment(horizontal="right"))

    TOTAL_ROW  = TEMPLATE_ROW + jumlah_items + 1
    total_semua = sum(it["jumlah"] * it["harga_jual"] for it in items)
    write_safe(ws, f"F{TOTAL_ROW}", "TOTAL",      Alignment(horizontal="right"))
    write_safe(ws, f"G{TOTAL_ROW}", total_semua,  Alignment(horizontal="right"))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"RAB_{nomor_faktur}_Blanko_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



# ============================================================
# === FAKTUR SEQ — nomor urut otomatis format INVDDMMYYNNN ===
# ============================================================
@app.route("/api/next_faktur_seq")
def next_faktur_seq():
    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM NotaBelanja")
    n = cur.fetchone()[0]
    conn.close()
    return jsonify({"next": n + 1})


# ============================================================
# === EXPORT FAKTUR PDF — transaksi_rab (invoice style) ===
# ============================================================
@app.route("/export_trx_pdf/<nomor_faktur>")
def export_trx_pdf(nomor_faktur):
    """Export faktur PDF bergaya invoice profesional seperti sampel INV."""
    import io
    from datetime import datetime as _dt
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur  = conn.cursor()

    cur.execute("SELECT * FROM PenjualanRAB WHERE nomor_faktur_beli = ?", (nomor_faktur,))
    pj = cur.fetchone()
    if not pj:
        conn.close()
        return f"Faktur {nomor_faktur} tidak ditemukan.", 404

    cur.execute("SELECT * FROM LayoutSetting LIMIT 1")
    setting = cur.fetchone()

    cur.execute("""
        SELECT nama_barang_jual AS nama, satuan_jual AS satuan,
               jumlah_jual AS qty, harga_jual AS harga
        FROM ItemPenjualanRAB WHERE penjualan_rab_id = ? ORDER BY id ASC
    """, (pj["id"],))
    raw = cur.fetchall()
    conn.close()

    if not raw:
        return f"Tidak ada item untuk faktur {nomor_faktur}.", 404

    # Grouping: nama+satuan sama dijumlah qty
    _g = {}
    for r in raw:
        if not r["nama"]: continue
        k = (str(r["nama"]).strip().lower(), str(r["satuan"] or "").strip().lower())
        if k not in _g:
            _g[k] = {"nama": r["nama"], "satuan": r["satuan"] or "",
                     "qty": float(r["qty"] or 0), "harga": float(r["harga"] or 0)}
        else:
            _g[k]["qty"]  += float(r["qty"] or 0)
            _g[k]["harga"] = float(r["harga"] or 0)
    items = list(_g.values())

    nama_kop = (setting["nama_koperasi"] if setting else None) or "KOPERASI KONSUMEN MARSUKI RAGIL MANDIRI"
    alamat   = (setting["alamat"]        if setting else None) or "Jl Ryacudu gg Surya Alam no 09 Korpri Raya, Kec. Sukarame Kota Bandar Lampung"
    kontak   = (setting["kontak"]        if setting else None) or "081262014034"

    tgl_jual = pj["tanggal_jual"] or ""
    try:    tgl_display = _dt.strptime(tgl_jual, "%Y-%m-%d").strftime("%d/%m/%Y")
    except: tgl_display = tgl_jual

    # ── Build PDF ─────────────────────────────────────────────
    buf    = io.BytesIO()
    W, _H  = A4
    mg     = 1.5 * cm
    TW     = W - 2 * mg   # ~510pt ≈ 18cm

    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=mg, rightMargin=mg,
                            topMargin=mg,  bottomMargin=mg)

    def PS(name, font="Helvetica", size=9, leading=13, align=TA_LEFT, bold=False):
        fn = "Helvetica-Bold" if bold else font
        return ParagraphStyle(name, fontName=fn, fontSize=size, leading=leading, alignment=align)

    content = []

    # ── HEADER: kop kiri | FAKTUR kanan ───────────────────────
    kop_p = Paragraph(
        f'<b>{nama_kop}</b><br/>{alamat}<br/>{kontak}',
        PS("kop", size=9, leading=14))

    faktur_p = Paragraph(
        f'<font size="22"><b>FAKTUR</b></font><br/>'
        f'<font size="11"><b>#{nomor_faktur}</b></font><br/>'
        f'<font size="9">Tanggal {tgl_display}</font>',
        PS("fak", size=9, leading=18, align=TA_RIGHT))

    hdr = Table([[kop_p, faktur_p]], colWidths=[TW * 0.60, TW * 0.40])
    hdr.setStyle(TableStyle([
        ("VALIGN",         (0,0),(-1,-1), "TOP"),
        ("LEFTPADDING",    (0,0),(-1,-1), 0),
        ("RIGHTPADDING",   (0,0),(-1,-1), 0),
        ("TOPPADDING",     (0,0),(-1,-1), 0),
        ("BOTTOMPADDING",  (0,0),(-1,-1), 6),
    ]))
    content.append(hdr)
    content.append(HRFlowable(width="100%", thickness=1.5, color=colors.black, spaceAfter=8))

    # ── TAGIH KEPADA ──────────────────────────────────────────
    content.append(Paragraph(f'<b>Tagih Kepada</b> {pj["konsumen"]}',
                              PS("tagih", size=10, leading=14)))
    content.append(Spacer(1, 10))

    # ── ITEMS TABLE ───────────────────────────────────────────
    # Cols: # | Item | Unit | Kuantitas | Biaya Satuan | Total
    # Widths sum = TW ≈ 510pt = 18cm
    cW = [0.65*cm, 6.1*cm, 1.9*cm, 2.2*cm, 3.1*cm, 4.05*cm]

    def ph(txt, align=TA_CENTER, bold=False):
        fn = "Helvetica-Bold" if bold else "Helvetica"
        return Paragraph(str(txt), ParagraphStyle("_", fontName=fn, fontSize=9,
                                                   leading=11, alignment=align))

    subtotal = 0
    tbl_data = [[
        ph("#",            bold=True),
        ph("Item",         TA_LEFT, bold=True),
        ph("Unit",         bold=True),
        ph("Kuantitas",    bold=True),
        ph("Biaya satuan", bold=True),
        ph("Total",        TA_RIGHT, bold=True),
    ]]
    for i, it in enumerate(items, 1):
        qty   = float(it["qty"]   or 0)
        harga = float(it["harga"] or 0)
        total = qty * harga
        subtotal += total
        qty_s = f"{qty:,.0f}" if qty == int(qty) else f"{qty:,.2f}"
        tbl_data.append([
            ph(i),
            ph(it["nama"],          TA_LEFT),
            ph(it["satuan"],        TA_CENTER),
            ph(qty_s,               TA_CENTER),
            ph(f"Rp{harga:,.2f}",   TA_RIGHT),
            ph(f"Rp{total:,.2f}",   TA_RIGHT),
        ])

    itbl = Table(tbl_data, colWidths=cW, repeatRows=1)
    itbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  colors.HexColor("#EEEEEE")),
        ("LINEABOVE",     (0,0), (-1,0),  1,    colors.black),
        ("LINEBELOW",     (0,0), (-1,0),  0.75, colors.black),
        ("LINEBELOW",     (0,-1),(-1,-1), 1,    colors.black),
        ("GRID",          (0,0), (-1,-1), 0.25, colors.HexColor("#CCCCCC")),
        ("BOX",           (0,0), (-1,-1), 0.75, colors.black),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, colors.HexColor("#FAFAFA")]),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("LEFTPADDING",   (0,0), (-1,-1), 4),
        ("RIGHTPADDING",  (0,0), (-1,-1), 4),
    ]))
    content.append(itbl)
    content.append(Spacer(1, 14))

    # ── BOTTOM: info bayar kiri | totals kanan ────────────────
    pay_lines = [
        Paragraph("<b>Info Pembayaran</b>", PS("pl", size=9, leading=13)),
        Spacer(1, 4),
        Paragraph("BRI 579601000052568",   PS("pb", size=8, leading=11)),
        Paragraph("An. Koperasi Konsumen Marsuki Ragil Mandiri",
                  PS("pb2", size=8, leading=11)),
        Spacer(1, 4),
        Paragraph("BSI 7307554806",        PS("pb3", size=8, leading=11)),
        Paragraph("An. Koperasi Konsumen Marsuki Ragil Mandiri",
                  PS("pb4", size=8, leading=11)),
    ]

    tot_str = f"Rp{subtotal:,.2f}"
    tot_tbl = Table(
        [["Subtotal",       tot_str],
         ["Total",          tot_str],
         ["Saldo Terutang", tot_str]],
        colWidths=[3.8*cm, 4.0*cm]
    )
    tot_tbl.setStyle(TableStyle([
        ("FONTNAME",     (0,0), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 9),
        ("ALIGN",        (0,0), (0,-1),  "LEFT"),
        ("ALIGN",        (1,0), (1,-1),  "RIGHT"),
        ("TOPPADDING",   (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0), (-1,-1), 3),
        ("LINEABOVE",    (0,0), (-1,0),  0.5, colors.black),
        ("LINEBELOW",    (0,-1),(-1,-1), 1.0, colors.black),
        ("BACKGROUND",   (0,-1),(-1,-1), colors.HexColor("#EEEEEE")),
    ]))

    bot = Table([[pay_lines, tot_tbl]],
                colWidths=[TW - 7.8*cm, 7.8*cm])
    bot.setStyle(TableStyle([
        ("VALIGN",       (0,0),(-1,-1), "TOP"),
        ("LEFTPADDING",  (0,0),(-1,-1), 0),
        ("RIGHTPADDING", (0,0),(-1,-1), 0),
        ("TOPPADDING",   (0,0),(-1,-1), 0),
        ("BOTTOMPADDING",(0,0),(-1,-1), 0),
    ]))
    content.append(bot)
    content.append(Spacer(1, 14))
    content.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey, spaceAfter=6))

    # ── FOOTER: kop + tanggal cetak (kanan) ───────────────────
    content.append(Paragraph(f"<b>{nama_kop}</b>",
                              PS("foot", size=9, leading=12, align=TA_RIGHT)))
    content.append(Paragraph(_dt.now().strftime("%d/%m/%Y"),
                              PS("footd", size=9, leading=12, align=TA_RIGHT)))

    doc.build(content)
    buf.seek(0)
    fname = f"Faktur_{nomor_faktur}_{_dt.now().strftime('%Y%m%d')}.pdf"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/pdf")


# ============================================================
# === EXPORT BLANKO XLSX (MasterBlanko.xlsx + Kop.png + stamp.png) ===
# ============================================================
@app.route("/export_trx_blanko/<nomor_faktur>")
def export_trx_blanko(nomor_faktur):
    import os, io
    from datetime import datetime
    from copy import copy as _copy
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.cell.cell import MergedCell
    base_path     = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_path, "MasterBlanko.xlsx")

    if not os.path.exists(template_path):
        return "❌ Template MasterBlanko.xlsx tidak ditemukan.", 500

    # ── Ambil data dari DB ────────────────────────────────────
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur  = conn.cursor()

    cur.execute("SELECT * FROM PenjualanRAB WHERE nomor_faktur_beli = ?", (nomor_faktur,))
    pj = cur.fetchone()
    if not pj:
        conn.close()
        return f"❌ Faktur RAB {nomor_faktur} tidak ditemukan.", 404

    cur.execute("""
        SELECT nama_barang_jual AS nama_barang,
               satuan_jual      AS satuan,
               jumlah_jual      AS jumlah,
               harga_jual
        FROM ItemPenjualanRAB
        WHERE penjualan_rab_id = ?
        ORDER BY id ASC
    """, (pj["id"],))
    rows = cur.fetchall()
    conn.close()

    if not rows:
        return f"❌ Tidak ada item untuk faktur {nomor_faktur}.", 404

    # Grouping: nama+satuan sama → jumlahkan qty
    _gb = {}
    for r in rows:
        if not r["nama_barang"]: continue
        k = (str(r["nama_barang"]).strip().lower(), str(r["satuan"] or "").strip().lower())
        if k not in _gb:
            _gb[k] = {"nama_barang": r["nama_barang"], "satuan": r["satuan"] or "",
                      "jumlah": float(r["jumlah"] or 0), "harga_jual": float(r["harga_jual"] or 0)}
        else:
            _gb[k]["jumlah"]    += float(r["jumlah"] or 0)
            _gb[k]["harga_jual"] = float(r["harga_jual"] or 0)
    items = list(_gb.values())
    N = len(items)

    # ── Load template ─────────────────────────────────────────
    # Struktur aktual MasterBlanko.xlsx:
    #   B7  = Tagih Kepada ...
    #   Row 9  = header kolom (B=#, C=Item, D=Unit, E=Kuantitas, F=Biaya satuan, G=Total)
    #   Row 10 = item 1  (G10=formula)
    #   Row 11 = item 2  (G11=formula)
    #   Row 12 = spacer
    #   Row 13 = Subtotal (F13), G13=SUM(G10:G11)
    #   Row 14 = Total    (F14), G14=G13
    #   Row 15 = Saldo Terutang (F15), G15=G14
    #   F23:G23 = nama koperasi (merged)
    #   F24:G24 = tanggal (merged)
    wb = load_workbook(template_path)
    ws = wb.active

    FIRST_ITEM_ROW = 10   # baris pertama item
    TEMPLATE_ITEMS = 2    # row 10 & 11
    INSERT_AT      = FIRST_ITEM_ROW + TEMPLATE_ITEMS  # = 12
    shift = max(0, N - TEMPLATE_ITEMS)

    # ── Insert baris tambahan jika item > 2 ──────────────────
    if shift > 0:
        from openpyxl.utils import get_column_letter

        # openpyxl TIDAK otomatis menggeser merged cells saat insert_rows.
        # Kumpulkan merged ranges yang berada di/bawah INSERT_AT, unmerge dulu,
        # insert, lalu re-merge di posisi yang sudah digeser.
        ranges_to_shift = [
            (rng.min_row, rng.min_col, rng.max_row, rng.max_col)
            for rng in list(ws.merged_cells.ranges)
            if rng.min_row >= INSERT_AT
        ]
        for r0, c0, r1, c1 in ranges_to_shift:
            ws.unmerge_cells(
                f"{get_column_letter(c0)}{r0}:{get_column_letter(c1)}{r1}"
            )

        ws.insert_rows(INSERT_AT, shift)

        for r0, c0, r1, c1 in ranges_to_shift:
            ws.merge_cells(
                f"{get_column_letter(c0)}{r0+shift}:{get_column_letter(c1)}{r1+shift}"
            )

        # Salin style dari template item row ke baris-baris baru
        for new_row in range(INSERT_AT, FIRST_ITEM_ROW + N):
            for col in range(2, 8):
                src = ws.cell(row=FIRST_ITEM_ROW, column=col)
                dst = ws.cell(row=new_row,        column=col)
                if src.has_style:
                    dst.font          = _copy(src.font)
                    dst.border        = _copy(src.border)
                    dst.fill          = _copy(src.fill)
                    dst.number_format = src.number_format
                    dst.alignment     = _copy(src.alignment)
            h = ws.row_dimensions[FIRST_ITEM_ROW].height
            if h: ws.row_dimensions[new_row].height = h

    # ── Isi data ──────────────────────────────────────────────
    ws["B7"].value = f"Tagih Kepada {pj['konsumen']}"

    for i, item in enumerate(items):
        r = FIRST_ITEM_ROW + i
        ws.cell(row=r, column=2).value = i + 1
        ws.cell(row=r, column=3).value = item["nama_barang"]
        ws.cell(row=r, column=4).value = item["satuan"]
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=5).value = item["jumlah"]
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=6).value = item["harga_jual"]
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=7).value = item["jumlah"] * item["harga_jual"]
        ws.cell(row=r, column=7).alignment = Alignment(horizontal="right")

    # Subtotal / Total / Saldo — originnya row 13/14/15, digeser oleh shift
    subtotal = sum(it["jumlah"] * it["harga_jual"] for it in items)
    ws.cell(row=13 + shift, column=7).value = subtotal
    ws.cell(row=14 + shift, column=7).value = subtotal
    ws.cell(row=15 + shift, column=7).value = subtotal

    # Koperasi name & tanggal — F23:G23 dan F24:G24 (otomatis shift oleh insert_rows)
    def _write_merged(ws, row, col, value, align=None):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            for rng in ws.merged_cells.ranges:
                if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                    top = ws.cell(row=rng.min_row, column=rng.min_col)
                    top.value = value
                    if align: top.alignment = align
                    return
        cell.value = value
        if align: cell.alignment = align

    _write_merged(ws, 23 + shift, 6, "Koperasi Konsumen Marsuki Ragil Mandiri",
                  Alignment(horizontal="center"))
    _write_merged(ws, 24 + shift, 6, datetime.now().strftime("%d/%m/%Y"),
                  Alignment(horizontal="center"))

    # ── Download xlsx ─────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"Blanko_{nomor_faktur}_{datetime.now().strftime('%d%m%Y')}.xlsx"
    return send_file(output, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ============================================================
# === SALIN MASTER.XLSX KE FOLDER REKAP ===
# ============================================================
@app.route("/api/salin_master/<nomor_faktur>")
def salin_master(nomor_faktur):
    import os, shutil
    from datetime import datetime as _dt, timedelta
    from copy import copy as _copy
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Alignment
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter

    base_path = os.path.dirname(os.path.abspath(__file__))
    src_path  = os.path.join(base_path, "Master.xlsx")
    dst_dir   = os.path.join(os.path.expanduser("~"), "Documents", "Rekap Nota Koperasi")

    if not os.path.exists(src_path):
        return jsonify({"status": "error", "message": "File Master.xlsx tidak ditemukan."})

    # ── Ambil data dari DB ────────────────────────────────────
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur  = conn.cursor()

    cur.execute("SELECT tanggal FROM NotaBelanja WHERE nomor_faktur = ?", (nomor_faktur,))
    nb = cur.fetchone()
    if not nb:
        conn.close()
        return jsonify({"status": "error", "message": f"Nota '{nomor_faktur}' tidak ditemukan."})

    cur.execute("SELECT * FROM PenjualanRAB WHERE nomor_faktur_beli = ?", (nomor_faktur,))
    pj = cur.fetchone()

    if pj:
        cur.execute("""
            SELECT nama_barang_jual AS nama, satuan_jual AS satuan,
                   jumlah_jual AS jumlah, harga_jual AS harga
            FROM ItemPenjualanRAB WHERE penjualan_rab_id = ? ORDER BY id ASC
        """, (pj["id"],))
        raw_items = cur.fetchall()
    else:
        raw_items = []
    conn.close()

    # Grouping nama+satuan sama
    _gb = {}
    for r in raw_items:
        if not r["nama"]: continue
        k = (str(r["nama"]).strip().lower(), str(r["satuan"] or "").strip().lower())
        if k not in _gb:
            _gb[k] = {"nama": r["nama"], "satuan": r["satuan"] or "",
                      "jumlah": float(r["jumlah"] or 0), "harga": float(r["harga"] or 0)}
        else:
            _gb[k]["jumlah"] += float(r["jumlah"] or 0)
            _gb[k]["harga"]   = float(r["harga"] or 0)
    items = list(_gb.values())
    N = len(items)

    try:
        tgl_obj    = _dt.strptime(nb["tanggal"], "%Y-%m-%d")
        tgl_minus1 = (tgl_obj - timedelta(days=1)).strftime("%d/%m/%Y")
        tgl_beli   = tgl_obj.strftime("%d/%m/%Y")
    except Exception:
        tgl_minus1 = tgl_beli = nb["tanggal"]

    # ── Salin & edit template ─────────────────────────────────
    os.makedirs(dst_dir, exist_ok=True)
    dst_path = os.path.join(dst_dir, f"{nomor_faktur}.xlsx")
    shutil.copy2(src_path, dst_path)

    wb = load_workbook(dst_path)
    ws = wb.active

    FIRST_ITEM_ROW = 10
    TEMPLATE_ITEMS = 2
    INSERT_AT      = FIRST_ITEM_ROW + TEMPLATE_ITEMS   # = 12
    shift = max(0, N - TEMPLATE_ITEMS)

    # ── Geser merged cells manual (openpyxl tidak otomatis) ──
    if shift > 0:
        ranges_to_shift = [
            (rng.min_row, rng.min_col, rng.max_row, rng.max_col)
            for rng in list(ws.merged_cells.ranges)
            if rng.min_row >= INSERT_AT
        ]
        for r0, c0, r1, c1 in ranges_to_shift:
            ws.unmerge_cells(f"{get_column_letter(c0)}{r0}:{get_column_letter(c1)}{r1}")

        ws.insert_rows(INSERT_AT, shift)

        for r0, c0, r1, c1 in ranges_to_shift:
            ws.merge_cells(f"{get_column_letter(c0)}{r0+shift}:{get_column_letter(c1)}{r1+shift}")

        # Salin style baris item ke baris baru
        for new_row in range(INSERT_AT, FIRST_ITEM_ROW + N):
            for col in range(2, 8):
                src = ws.cell(row=FIRST_ITEM_ROW, column=col)
                dst = ws.cell(row=new_row,        column=col)
                if src.has_style:
                    dst.font          = _copy(src.font)
                    dst.border        = _copy(src.border)
                    dst.fill          = _copy(src.fill)
                    dst.number_format = src.number_format
                    dst.alignment     = _copy(src.alignment)
            h = ws.row_dimensions[FIRST_ITEM_ROW].height
            if h: ws.row_dimensions[new_row].height = h

    # ── Isi header ────────────────────────────────────────────
    ws["G4"].value = f"#{nomor_faktur}"
    ws["G5"].value = f"Tanggal {tgl_minus1}"
    ws["B7"].value = f"Tagih Kepada {pj['konsumen']}" if pj else ws["B7"].value

    # ── Isi baris item ────────────────────────────────────────
    for i, item in enumerate(items):
        r = FIRST_ITEM_ROW + i
        ws.cell(row=r, column=2).value = i + 1
        ws.cell(row=r, column=3).value = item["nama"]
        ws.cell(row=r, column=4).value = item["satuan"]
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=5).value = item["jumlah"]
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=6).value = item["harga"]
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=7).value = item["jumlah"] * item["harga"]
        ws.cell(row=r, column=7).alignment = Alignment(horizontal="right")

    # ── Subtotal / Total / Saldo ──────────────────────────────
    subtotal = sum(it["jumlah"] * it["harga"] for it in items)
    saldo_row = 15 + shift
    ws.cell(row=13 + shift, column=7).value = subtotal
    ws.cell(row=14 + shift, column=7).value = subtotal
    ws.cell(row=saldo_row,  column=7).value = subtotal

    # ── Tanggal beli di F24 (shifted) ────────────────────────
    ws.cell(row=24 + shift, column=6).value = tgl_beli

    # ── Background #D0CECE pada header tabel B9:G9 ───────────
    header_fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
    for col in range(2, 8):
        ws.cell(row=9, column=col).fill = header_fill

    wb.save(dst_path)

    return jsonify({"status": "ok",
                    "message": f"File disalin ke {dst_path}"})


# ============================================================
# === ESTIMASI BAHAN ===
# ============================================================

@app.route("/estimasi_bahan")
def estimasi_bahan():
    return render_template("estimasi_bahan.html", today=date.today().strftime("%Y-%m-%d"))


@app.route("/api/estimasi_bahan/simpan", methods=["POST"])
def simpan_estimasi_bahan():
    data = request.get_json()
    if not data:
        return jsonify({"status": "error", "message": "Data kosong"})

    tanggal        = data.get("tanggal", "").strip()
    nama_menu      = data.get("nama_menu", "").strip()
    nomor_estimasi = data.get("nomor_estimasi", "").strip()
    items          = data.get("items", [])

    if not tanggal or not nama_menu or not nomor_estimasi:
        return jsonify({"status": "error", "message": "Tanggal, nama menu, dan nomor estimasi wajib diisi."})
    if not items:
        return jsonify({"status": "error", "message": "Minimal 1 item harus diisi."})
    if len(nama_menu) > 150:
        return jsonify({"status": "error", "message": "Nama menu maksimal 150 karakter."})

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    try:
        # Hapus lama jika nomor_estimasi sama
        cur.execute("SELECT id FROM EstimasiBahan WHERE nomor_estimasi = ?", (nomor_estimasi,))
        existing = cur.fetchone()
        if existing:
            cur.execute("DELETE FROM ItemEstimasiBahan WHERE estimasi_id = ?", (existing[0],))
            cur.execute("DELETE FROM EstimasiBahan WHERE id = ?", (existing[0],))

        cur.execute("""
            INSERT INTO EstimasiBahan (tanggal, nama_menu, nomor_estimasi)
            VALUES (?, ?, ?)
        """, (tanggal, nama_menu, nomor_estimasi))
        est_id = cur.lastrowid

        saved = 0
        for it in items:
            nama   = str(it.get("nama", "")).strip()
            jumlah = float(it.get("jumlah", 0) or 0)
            satuan = str(it.get("satuan", "")).strip()
            harga  = float(it.get("harga", 0) or 0)
            if not nama or harga <= 0:
                continue
            total = jumlah * harga
            cur.execute("""
                INSERT INTO ItemEstimasiBahan (estimasi_id, nama_barang, jumlah, satuan, harga_satuan, total_harga)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (est_id, nama, jumlah, satuan, harga, total))
            saved += 1

        if saved == 0:
            conn.rollback()
            return jsonify({"status": "error", "message": "Tidak ada item valid yang tersimpan."})

        conn.commit()
        return jsonify({"status": "ok", "message": f"{saved} item berhasil disimpan.", "id": est_id})
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": str(e)})
    finally:
        conn.close()


@app.route("/api/estimasi_bahan/list")
def list_estimasi_bahan():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT e.id, e.tanggal, e.nama_menu, e.nomor_estimasi, e.dibuat_pada,
               COUNT(i.id) as jumlah_item,
               SUM(i.total_harga) as grand_total
        FROM EstimasiBahan e
        LEFT JOIN ItemEstimasiBahan i ON e.id = i.estimasi_id
        GROUP BY e.id
        ORDER BY e.id DESC LIMIT 100
    """)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"status": "ok", "data": rows})


@app.route("/api/estimasi_bahan/get/<nomor_estimasi>")
def get_estimasi_bahan(nomor_estimasi):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM EstimasiBahan WHERE nomor_estimasi = ?", (nomor_estimasi,))
    header = cur.fetchone()
    if not header:
        conn.close()
        return jsonify({"status": "not_found"})
    cur.execute("SELECT * FROM ItemEstimasiBahan WHERE estimasi_id = ?", (header["id"],))
    items = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"status": "ok", "header": dict(header), "items": items})


@app.route("/api/estimasi_bahan/hapus/<int:est_id>", methods=["POST"])
def hapus_estimasi_bahan(est_id):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("DELETE FROM ItemEstimasiBahan WHERE estimasi_id = ?", (est_id,))
    cur.execute("DELETE FROM EstimasiBahan WHERE id = ?", (est_id,))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok"})


@app.route("/export_estimasi_excel")
def export_estimasi_excel():
    tanggal = request.args.get("tanggal", "").strip()
    nomor   = request.args.get("nomor_estimasi", "").strip()

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if nomor:
        cur.execute("""
            SELECT e.tanggal, e.nama_menu, e.nomor_estimasi,
                   i.nama_barang, i.jumlah, i.satuan, i.harga_satuan, i.total_harga
            FROM EstimasiBahan e
            JOIN ItemEstimasiBahan i ON e.id = i.estimasi_id
            WHERE e.nomor_estimasi = ?
            ORDER BY i.id ASC
        """, (nomor,))
    elif tanggal:
        cur.execute("""
            SELECT e.tanggal, e.nama_menu, e.nomor_estimasi,
                   i.nama_barang, i.jumlah, i.satuan, i.harga_satuan, i.total_harga
            FROM EstimasiBahan e
            JOIN ItemEstimasiBahan i ON e.id = i.estimasi_id
            WHERE e.tanggal = ?
            ORDER BY e.nama_menu ASC, i.id ASC
        """, (tanggal,))
    else:
        conn.close()
        return jsonify({"status": "error", "message": "Parameter tanggal atau nomor_estimasi diperlukan."}), 400

    rows = cur.fetchall()
    conn.close()

    if not rows:
        return "Tidak ada data estimasi untuk diekspor.", 404

    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        wb  = writer.book
        ws  = wb.add_worksheet("Estimasi Bahan")
        writer.sheets["Estimasi Bahan"] = ws

        # Formats
        title_fmt = wb.add_format({"bold": True, "font_size": 14, "align": "center", "valign": "vcenter"})
        info_fmt  = wb.add_format({"bold": True})
        hdr_fmt   = wb.add_format({"bold": True, "align": "center", "valign": "vcenter",
                                    "bg_color": "#B7DEE8", "border": 1})
        text_fmt  = wb.add_format({"border": 1})
        money_fmt = wb.add_format({"num_format": "#,##0", "align": "right", "border": 1})
        total_lbl = wb.add_format({"bold": True, "align": "right", "bg_color": "#E2EFDA", "border": 1})
        total_val = wb.add_format({"bold": True, "num_format": "#,##0",
                                    "align": "right", "bg_color": "#E2EFDA", "border": 1})
        menu_hdr  = wb.add_format({"bold": True, "font_size": 11, "bg_color": "#FFF2CC",
                                    "border": 1, "valign": "vcenter"})

        # Judul
        ws.merge_range("A1:H1", "LAPORAN ESTIMASI BAHAN", title_fmt)
        ws.set_row(0, 24)

        # Info
        if tanggal and not nomor:
            ws.write("A3", "Tanggal :", info_fmt); ws.write("B3", tanggal)
        else:
            ws.write("A3", "No. Estimasi :", info_fmt); ws.write("B3", nomor)

        ws.write("A4", "Diekspor :", info_fmt)
        ws.write("B4", datetime.now().strftime("%d/%m/%Y %H:%M"))

        # Lebar kolom
        ws.set_column(0, 0, 12)  # Tanggal
        ws.set_column(1, 1, 40)  # Nama Menu
        ws.set_column(2, 2, 18)  # No. Estimasi
        ws.set_column(3, 3, 28)  # Nama Barang
        ws.set_column(4, 4, 10)  # Jumlah
        ws.set_column(5, 5, 10)  # Satuan
        ws.set_column(6, 6, 15)  # Harga Satuan
        ws.set_column(7, 7, 15)  # Total

        # Header kolom
        headers = ["Tanggal", "Nama Menu", "No. Estimasi", "Nama Barang", "Jumlah", "Satuan", "Harga Satuan", "Total"]
        for ci, h in enumerate(headers):
            ws.write(5, ci, h, hdr_fmt)

        # Group per menu agar lebih rapi
        from collections import defaultdict
        grouped = defaultdict(list)
        for r in rows:
            grouped[(r["tanggal"], r["nama_menu"], r["nomor_estimasi"])].append(r)

        data_row = 6
        grand_total = 0
        for (tgl, menu, no_est), items in grouped.items():
            menu_subtotal = 0
            for r in items:
                ws.write(data_row, 0, tgl, text_fmt)
                ws.write(data_row, 1, menu, text_fmt)
                ws.write(data_row, 2, no_est, text_fmt)
                ws.write(data_row, 3, r["nama_barang"], text_fmt)
                ws.write(data_row, 4, r["jumlah"], money_fmt)
                ws.write(data_row, 5, r["satuan"], text_fmt)
                ws.write(data_row, 6, r["harga_satuan"], money_fmt)
                ws.write(data_row, 7, r["total_harga"], money_fmt)
                menu_subtotal += (r["total_harga"] or 0)
                data_row += 1

            # Subtotal per menu
            ws.write(data_row, 6, f"Subtotal {menu[:30]}:", total_lbl)
            ws.write(data_row, 7, menu_subtotal, total_val)
            data_row += 1
            grand_total += menu_subtotal

        # Grand total
        ws.write(data_row + 1, 6, "GRAND TOTAL:", total_lbl)
        ws.write(data_row + 1, 7, grand_total, total_val)

    output.seek(0)
    if nomor:
        fname = f"Estimasi_{nomor}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    else:
        fname = f"Estimasi_{tanggal}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(output, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    init_db()
    app.run(
        host="0.0.0.0",  # ⬅️ WAJIB agar bisa diakses intranet
        port=5000,
        debug=False     # ⬅️ disarankan untuk intranet
    )

